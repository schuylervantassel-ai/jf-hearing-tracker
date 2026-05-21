#!/usr/bin/env python3
"""
Jamestown Foundation Congressional Hearing Tracker — Web UI
Run:  python app.py
Open: http://localhost:5000
"""

import io
import os
import shutil
import sys
import threading
import webbrowser
from datetime import date, datetime, timezone

# Background RSS/API/social import: set RSS_AUTO_PULL=0 to disable.
# RSS_PULL_INTERVAL_MINUTES defaults to 60 (minimum 1).
def _env_truthy(name: str, default: str = "1") -> bool:
    return os.environ.get(name, default).strip().lower() not in ("0", "false", "no", "")

RSS_AUTO_PULL_ENABLED = _env_truthy("RSS_AUTO_PULL", "1")
RSS_POLL_INTERVAL_MIN = max(1, int(os.environ.get("RSS_PULL_INTERVAL_MINUTES", "60")))
# Seconds to wait before the first auto-pull (subsequent waits use the interval above).
RSS_AUTO_PULL_INITIAL_DELAY_SEC = max(0, int(os.environ.get("RSS_AUTO_PULL_INITIAL_DELAY_SEC", "10")))
FR_AUTO_PULL_ENABLED = _env_truthy("FR_AUTO_PULL", "0")
FR_PULL_INTERVAL_MIN = max(60, int(os.environ.get("FR_PULL_INTERVAL_MINUTES", "360")))
DATA_LOCK = threading.Lock()
_last_pull: dict = {
    "ts": None,
    "new_rss": 0,
    "new_api": 0,
    "new_social": 0,
    "api_updated": 0,
    "in_progress": False,
    "error": None,
    "message": None,
}

from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, url_for

# ── Path resolution (works both in dev and when bundled with PyInstaller) ─────
def _base_dir():
    """Root directory for bundled resources (templates) and data files."""
    if getattr(sys, "frozen", False):
        # Running inside a PyInstaller bundle
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = _base_dir()
sys.path.insert(0, BASE_DIR)


def _data_root() -> str:
    return os.environ.get("DATA_DIR", "").strip() or BASE_DIR


def _seed_persistent_data() -> None:
    """Copy bundled JSON into DATA_DIR on first deploy (Render persistent disk)."""
    data_dir = os.environ.get("DATA_DIR", "").strip()
    if not data_dir:
        return
    os.makedirs(data_dir, exist_ok=True)
    for name in (
        "hearings.json",
        "rss_config.json",
        "social_config.json",
        "social_feed.json",
        "congress_api.json",
        "congress_config.json",
        "govtrack_committees.json",
        "chamber_calendar.json",
        "fr_watchlist.json",
        "fr_documents.json",
    ):
        dest = os.path.join(data_dir, name)
        src = os.path.join(BASE_DIR, name)
        if not os.path.exists(dest) and os.path.exists(src):
            shutil.copy2(src, dest)


_seed_persistent_data()

from comit import (
    ACTIONS, COMMITTEES, JAMESTOWN_ANGLES, STATUS_OPTIONS,
    load_data, load_feeds, next_id, pull_rss_feeds, pull_congress_api,
    save_data, save_feeds,
    load_social_feeds, save_social_feeds, load_social_items, pull_social_feeds,
    build_heatmap_points, DC_BUILDINGS,
    load_fr_documents, save_fr_documents, load_fr_watchlist, save_fr_watchlist,
    pull_federal_register, fr_comment_period_label, FR_WORKFLOW_STATUSES,
    load_congress_config,
    pull_senate_schedule, refresh_govtrack_committee_cache,
    build_committee_ongoings, load_govtrack_cache,
    pull_session_calendar, build_session_calendar_week, build_session_calendar_month,
    load_chamber_calendar,
)
import json as _json

# Point Flask at the correct templates folder when bundled
_template_dir = os.path.join(
    getattr(sys, "_MEIPASS", BASE_DIR), "templates"
)
app = Flask(__name__, template_folder=_template_dir)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "jf-hearing-tracker-dev-only")

# ── Badge helpers (full class strings so Tailwind CDN picks them up) ──────────

def status_cls(status):
    return {
        "Upcoming":  "bg-blue-100 text-blue-800",
        "Completed": "bg-emerald-100 text-emerald-800",
        "Cancelled": "bg-red-100 text-red-800",
        "Postponed": "bg-amber-100 text-amber-800",
    }.get(status, "bg-slate-100 text-slate-700")

def angle_cls(angle):
    return {
        "Russia/Eurasia":            "bg-red-100 text-red-700",
        "China/Indo-Pacific":        "bg-yellow-100 text-yellow-700",
        "Middle East/North Africa":  "bg-orange-100 text-orange-700",
        "Sub-Saharan Africa":        "bg-emerald-100 text-emerald-700",
        "Terrorism/Extremism":       "bg-rose-100 text-rose-700",
        "Central Asia":              "bg-purple-100 text-purple-700",
        "Latin America":             "bg-teal-100 text-teal-700",
        "Cyber/Information Warfare": "bg-sky-100 text-sky-700",
        "Other":                     "bg-slate-100 text-slate-600",
    }.get(angle, "bg-slate-100 text-slate-600")

def action_cls(action):
    return {
        "Send brief":       "bg-orange-100 text-orange-700",
        "Send questions":   "bg-orange-100 text-orange-700",
        "Offer testimony":  "bg-red-100 text-red-700",
        "Request meeting":  "bg-blue-100 text-blue-700",
        "Monitor only":     "bg-slate-100 text-slate-600",
        "No action needed": "bg-emerald-100 text-emerald-700",
    }.get(action, "bg-slate-100 text-slate-600")

def fr_workflow_cls(status):
    return {
        "Watching":        "bg-slate-100 text-slate-600",
        "Drafting":        "bg-amber-100 text-amber-800",
        "Ready to file":   "bg-blue-100 text-blue-800",
        "Filed":           "bg-emerald-100 text-emerald-800",
        "Closed":          "bg-slate-200 text-slate-500",
    }.get(status, "bg-slate-100 text-slate-600")

def fr_period_cls(comments_close_on):
    label = fr_comment_period_label(comments_close_on)
    if label == "Closed":
        return "bg-slate-100 text-slate-500"
    if label in ("Closes today",) or label.startswith("Closes in"):
        return "bg-red-100 text-red-800"
    if label == "Open":
        return "bg-emerald-100 text-emerald-800"
    return "bg-slate-100 text-slate-500"

def fr_period_label(comments_close_on):
    return fr_comment_period_label(comments_close_on)

def days_away(date_str):
    try:
        return (date.fromisoformat(date_str) - date.today()).days
    except Exception:
        return None

MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

app.jinja_env.globals.update(
    status_cls=status_cls,
    angle_cls=angle_cls,
    action_cls=action_cls,
    fr_workflow_cls=fr_workflow_cls,
    fr_period_cls=fr_period_cls,
    fr_period_label=fr_period_label,
    days_away=days_away,
    today=date.today,
    abs=abs,
    MONTH_ABBR=MONTH_ABBR,
)

# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route("/")
def dashboard():
    hearings = load_data()
    feeds    = load_feeds()
    today_d  = date.today()

    upcoming_14 = sorted(
        [h for h in hearings
         if h.get("status") in ("Upcoming", None)
         and 0 <= (date.fromisoformat(h["date"]) - today_d).days <= 14],
        key=lambda x: x["date"],
    )

    needs_action = [
        h for h in hearings
        if h.get("action") in ("Send brief", "Send questions", "Offer testimony", "Request meeting")
        and h.get("status") == "Upcoming"
    ]

    active_feeds = sum(1 for f in feeds if f.get("active", True))

    comm_counts = {}
    for h in hearings:
        c = h.get("committee", "Other")
        comm_counts[c] = comm_counts.get(c, 0) + 1

    angle_counts = {}
    for h in hearings:
        a = h.get("angle", "Other")
        angle_counts[a] = angle_counts.get(a, 0) + 1

    return render_template(
        "dashboard.html",
        total=len(hearings),
        upcoming_14=upcoming_14,
        needs_action=needs_action,
        active_feeds=active_feeds,
        comm_counts=sorted(comm_counts.items(), key=lambda x: -x[1]),
        angle_counts=sorted(angle_counts.items(), key=lambda x: -x[1]),
        committees=COMMITTEES + ["Multiple"],
    )


@app.route("/hearings")
def hearings_list():
    hearings = load_data()
    q  = request.args.get("q", "").lower()
    sf = request.args.get("status", "")
    cf = request.args.get("committee", "")
    af = request.args.get("angle", "")

    filtered = hearings
    if q:
        filtered = [h for h in filtered if any(
            q in h.get(f, "").lower()
            for f in ["topic", "committee", "angle", "witnesses", "notes"]
        )]
    if sf:
        filtered = [h for h in filtered if h.get("status") == sf]
    if cf:
        filtered = [h for h in filtered if h.get("committee") == cf]
    if af:
        filtered = [h for h in filtered if h.get("angle") == af]

    filtered.sort(key=lambda x: x["date"], reverse=True)

    return render_template(
        "hearings.html",
        hearings=filtered,
        total_all=len(hearings),
        committees=COMMITTEES,
        angles=JAMESTOWN_ANGLES,
        statuses=STATUS_OPTIONS,
        q=request.args.get("q", ""),
        sf=sf, cf=cf, af=af,
    )


@app.route("/hearing/<int:hid>")
def hearing_detail(hid):
    hearings = load_data()
    h = next((x for x in hearings if x["id"] == hid), None)
    if not h:
        flash("Hearing not found.", "error")
        return redirect(url_for("hearings_list"))
    return render_template("detail.html", h=h)


@app.route("/add", methods=["GET", "POST"])
def add_hearing():
    if request.method == "POST":
        with DATA_LOCK:
            hearings = load_data()
            h = {
                "id":        next_id(hearings),
                "date":      request.form["date"],
                "committee": request.form["committee"],
                "topic":     request.form["topic"],
                "witnesses": request.form.get("witnesses", ""),
                "angle":     request.form["angle"],
                "action":    request.form["action"],
                "status":    request.form["status"],
                "url":       request.form.get("url", ""),
                "questions": request.form.get("questions", "").strip().replace("\n", " | "),
                "notes":     request.form.get("notes", ""),
                "source":    "manual",
                "created":   datetime.now().strftime("%Y-%m-%d"),
            }
            hearings.append(h)
            save_data(hearings)
        flash(f"Hearing #{h['id']:03d} added.", "success")
        return redirect(url_for("hearing_detail", hid=h["id"]))

    return render_template(
        "form.html", mode="add", h={},
        committees=COMMITTEES, angles=JAMESTOWN_ANGLES,
        actions=ACTIONS, statuses=STATUS_OPTIONS,
        default_date=date.today().isoformat(),
    )


@app.route("/edit/<int:hid>", methods=["GET", "POST"])
def edit_hearing(hid):
    if request.method == "POST":
        with DATA_LOCK:
            hearings = load_data()
            h = next((x for x in hearings if x["id"] == hid), None)
            if h:
                h["date"]      = request.form["date"]
                h["committee"] = request.form["committee"]
                h["topic"]     = request.form["topic"]
                h["witnesses"] = request.form.get("witnesses", "")
                h["angle"]     = request.form["angle"]
                h["action"]    = request.form["action"]
                h["status"]    = request.form["status"]
                h["url"]       = request.form.get("url", "")
                h["questions"] = request.form.get("questions", "").strip().replace("\n", " | ")
                h["notes"]     = request.form.get("notes", "")
                save_data(hearings)
        if not h:
            flash("Hearing not found.", "error")
            return redirect(url_for("hearings_list"))
        flash(f"Hearing #{h['id']:03d} updated.", "success")
        return redirect(url_for("hearing_detail", hid=hid))

    hearings = load_data()
    h = next((x for x in hearings if x["id"] == hid), None)
    if not h:
        flash("Hearing not found.", "error")
        return redirect(url_for("hearings_list"))

    h_form = dict(h)
    h_form["questions"] = h.get("questions", "").replace(" | ", "\n")
    return render_template(
        "form.html", mode="edit", h=h_form,
        committees=COMMITTEES, angles=JAMESTOWN_ANGLES,
        actions=ACTIONS, statuses=STATUS_OPTIONS,
    )


# ── Federal Register ──────────────────────────────────────────────────────────

def _fr_filter_documents(documents, q, wf, cf):
    filtered = documents
    if q:
        filtered = [
            d for d in filtered
            if q in (d.get("title") or "").lower()
            or q in (d.get("abstract") or "").lower()
            or q in " ".join(d.get("agencies") or []).lower()
            or q in (d.get("document_number") or "").lower()
            or any(q in (x or "").lower() for x in (d.get("docket_ids") or []))
        ]
    if wf:
        filtered = [d for d in filtered if d.get("workflow_status") == wf]
    today = date.today()
    if cf == "open":
        filtered = [
            d for d in filtered
            if d.get("comments_close_on")
            and d["comments_close_on"] >= today.isoformat()
        ]
    elif cf == "soon":
        filtered = [
            d for d in filtered
            if d.get("comments_close_on")
            and 0 <= (date.fromisoformat(d["comments_close_on"][:10]) - today).days <= 14
        ]
    elif cf == "closed":
        filtered = [
            d for d in filtered
            if d.get("comments_close_on")
            and d["comments_close_on"] < today.isoformat()
        ]
    filtered.sort(
        key=lambda x: (x.get("publication_date") or "", x.get("document_number") or ""),
        reverse=True,
    )
    return filtered


@app.route("/federal-register")
def federal_register_list():
    documents = load_fr_documents()
    q = request.args.get("q", "").lower().strip()
    wf = request.args.get("workflow", "")
    cf = request.args.get("comments", "")
    filtered = _fr_filter_documents(documents, q, wf, cf)
    return render_template(
        "federal_register.html",
        documents=filtered,
        total_all=len(documents),
        watchlist=load_fr_watchlist(),
        workflow_statuses=FR_WORKFLOW_STATUSES,
        fr_api_key_configured=bool(_load_fr_api_key()),
        q=request.args.get("q", ""),
        wf=wf,
        cf=cf,
    )


@app.route("/federal-register/pull", methods=["POST"])
def fr_pull():
    with DATA_LOCK:
        documents = load_fr_documents()
        watchlist = load_fr_watchlist()
        new_items = pull_federal_register(documents, watchlist, silent=True)
    flash(f"{len(new_items)} new Federal Register document(s) imported.", "success")
    return redirect(url_for("federal_register_list"))


@app.route("/federal-register/<int:did>")
def fr_detail(did):
    documents = load_fr_documents()
    d = next((x for x in documents if x["id"] == did), None)
    if not d:
        flash("Document not found.", "error")
        return redirect(url_for("federal_register_list"))
    return render_template(
        "fr_detail.html",
        d=d,
        workflow_statuses=FR_WORKFLOW_STATUSES,
    )


@app.route("/federal-register/<int:did>/save", methods=["POST"])
def fr_save(did):
    with DATA_LOCK:
        documents = load_fr_documents()
        d = next((x for x in documents if x["id"] == did), None)
        if not d:
            flash("Document not found.", "error")
            return redirect(url_for("federal_register_list"))
        if "workflow_status" in request.form:
            ws = request.form.get("workflow_status")
            if ws in FR_WORKFLOW_STATUSES:
                d["workflow_status"] = ws
        if "draft_comment" in request.form:
            d["draft_comment"] = request.form.get("draft_comment", "")
        if "notes" in request.form:
            d["notes"] = request.form.get("notes", "")
        save_fr_documents(documents)
    flash("Document updated.", "success")
    return redirect(url_for("fr_detail", did=did))


@app.route("/federal-register/<int:did>/delete", methods=["POST"])
def fr_delete(did):
    with DATA_LOCK:
        documents = [x for x in load_fr_documents() if x["id"] != did]
        save_fr_documents(documents)
    flash("Document removed.", "info")
    return redirect(url_for("federal_register_list"))


@app.route("/federal-register/watch/toggle/<int:idx>", methods=["POST"])
def fr_toggle_watch(idx):
    with DATA_LOCK:
        watchlist = load_fr_watchlist()
        if 0 <= idx < len(watchlist):
            watchlist[idx]["active"] = not watchlist[idx].get("active", True)
            save_fr_watchlist(watchlist)
    return redirect(url_for("federal_register_list"))


@app.route("/federal-register/watch/add", methods=["POST"])
def fr_add_watch():
    with DATA_LOCK:
        watchlist = load_fr_watchlist()
        watchlist.append({
            "name": request.form["name"].strip(),
            "search_term": request.form["search_term"].strip(),
            "document_types": ["PRORULE", "NOTICE"],
            "active": True,
        })
        save_fr_watchlist(watchlist)
    flash(f"Watch “{request.form['name'].strip()}” added.", "success")
    return redirect(url_for("federal_register_list"))


@app.route("/delete/<int:hid>", methods=["POST"])
def delete_hearing(hid):
    with DATA_LOCK:
        hearings = [h for h in load_data() if h["id"] != hid]
        save_data(hearings)
    flash("Hearing deleted.", "info")
    return redirect(url_for("hearings_list"))


_API_KEY_FILE = os.path.join(_data_root(), "congress_api.json")
_FR_API_KEY_FILE = os.path.join(_data_root(), "federal_register_api.json")

def _load_fr_api_key():
    key = os.environ.get("FEDERAL_REGISTER_API_KEY", "").strip()
    if key:
        return key
    try:
        with open(_FR_API_KEY_FILE) as f:
            return _json.load(f).get("api_key", "")
    except Exception:
        return ""

def _load_api_key():
    key = os.environ.get("CONGRESS_API_KEY", "").strip()
    if key:
        return key
    try:
        with open(_API_KEY_FILE) as f:
            return _json.load(f).get("api_key", "")
    except Exception:
        return ""


def _mask_api_key(key: str) -> str:
    """Display hint only — never show full key in HTML."""
    key = (key or "").strip()
    if not key:
        return ""
    if len(key) <= 4:
        return "••••"
    return "•" * (len(key) - 4) + key[-4:]


def _save_api_key(key):
    with open(_API_KEY_FILE, "w") as f:
        _json.dump({"api_key": key}, f)


def _run_hearing_feed_import():
    """
    RSS + Senate hearing schedule + Congress.gov API + social.
    Kept separate from session-calendar / GovTrack pulls (those are slow).
    """
    hearings = load_data()
    feeds = load_feeds()
    social_feeds = load_social_feeds()
    new_rss = pull_rss_feeds(hearings, feeds, silent=True)
    sched = pull_senate_schedule(hearings, silent=True)
    api_key = _load_api_key()
    api_result = (
        pull_congress_api(hearings, api_key, silent=True)
        if api_key
        else {"new": [], "updated": 0}
    )
    new_api = api_result.get("new", [])
    new_social = pull_social_feeds(social_feeds, silent=True)
    return {
        "new_rss": new_rss,
        "new_api": new_api,
        "new_social": new_social,
        "api_updated": api_result.get("updated", 0),
        "sched_new": len(sched.get("new", [])),
        "sched_upd": sched.get("updated", 0),
    }


def _format_feed_pull_message(result):
    parts = []
    if result.get("new_rss"):
        parts.append(f"{len(result['new_rss'])} RSS")
    if result.get("sched_new"):
        parts.append(f"{result['sched_new']} Senate schedule")
    if result.get("sched_upd"):
        parts.append(f"{result['sched_upd']} schedule updates")
    if result.get("new_api"):
        parts.append(f"{len(result['new_api'])} Congress.gov")
    if result.get("api_updated"):
        parts.append(f"{result['api_updated']} API updates")
    if result.get("new_social"):
        parts.append(f"{len(result['new_social'])} social")
    if parts:
        return "Feed refresh finished: " + ", ".join(parts) + "."
    return "Feed refresh finished (no new items)."


def _feed_pull_worker():
    """Background job — must not block the Gunicorn worker."""
    try:
        with DATA_LOCK:
            result = _run_hearing_feed_import()
        _last_pull["ts"] = datetime.now(timezone.utc).isoformat()
        _last_pull["new_rss"] = len(result["new_rss"])
        _last_pull["new_api"] = len(result["new_api"])
        _last_pull["new_social"] = len(result["new_social"])
        _last_pull["api_updated"] = result["api_updated"]
        _last_pull["error"] = None
        _last_pull["message"] = _format_feed_pull_message(result)
        print(f"[feed-pull] {_last_pull['message']}", flush=True)
    except Exception as e:
        _last_pull["error"] = str(e)
        _last_pull["message"] = f"Feed refresh failed: {e}"
        print(f"[feed-pull] {e}", file=sys.stderr, flush=True)
    finally:
        _last_pull["in_progress"] = False


def _start_background_feed_pull():
    """Return False if a pull is already running."""
    if _last_pull.get("in_progress"):
        return False
    _last_pull["in_progress"] = True
    _last_pull["error"] = None
    _last_pull["message"] = None
    threading.Thread(target=_feed_pull_worker, name="FeedPull", daemon=True).start()
    return True


def _auto_feed_poll_loop():
    """Daemon: sleep between full imports while the server process runs."""
    import time

    interval_sec = RSS_POLL_INTERVAL_MIN * 60
    first_wait = True
    while True:
        wait = (
            min(RSS_AUTO_PULL_INITIAL_DELAY_SEC, interval_sec)
            if first_wait
            else interval_sec
        )
        first_wait = False
        time.sleep(wait)
        if not RSS_AUTO_PULL_ENABLED:
            continue
        if not _start_background_feed_pull():
            print("[RSS auto-pull] skipped (pull already in progress)", flush=True)


def _start_auto_feed_poller():
    if not RSS_AUTO_PULL_ENABLED:
        return
    threading.Thread(
        target=_auto_feed_poll_loop, name="RSSAutoPull", daemon=True
    ).start()


@app.route("/api/poll-status")
def poll_status():
    return jsonify(_last_pull)


def _flash_pull_started(redirect_url):
    flash(
        "Feed refresh started in the background (usually 1–2 minutes). "
        "You can keep using the site; check back shortly.",
        "info",
    )
    return redirect(redirect_url)


@app.route("/feeds")
def feeds_page():
    ccfg = load_congress_config()
    return render_template(
        "feeds.html",
        feeds=load_feeds(),
        committees=COMMITTEES + ["Multiple"],
        api_key_configured=bool(_load_api_key()),
        api_key_masked=_mask_api_key(_load_api_key()),
        congress_config=ccfg,
        rss_auto_pull_enabled=RSS_AUTO_PULL_ENABLED,
        rss_poll_interval_min=RSS_POLL_INTERVAL_MIN,
        last_pull=_last_pull,
    )


@app.route("/feeds/pull", methods=["POST"])
def pull_feeds():
    if not _start_background_feed_pull():
        flash("A feed refresh is already running. Please wait a minute.", "error")
        return redirect(url_for("feeds_page"))
    return _flash_pull_started(url_for("feeds_page"))


@app.route("/feeds/pull-api", methods=["POST"])
def pull_api_only():
    api_key = _load_api_key()
    if not api_key:
        flash("Add your Congress.gov API key first.", "error")
        return redirect(url_for("feeds_page"))
    if _last_pull.get("in_progress"):
        flash("A feed refresh is already running. Please wait.", "error")
        return redirect(url_for("feeds_page"))

    def _api_worker():
        try:
            with DATA_LOCK:
                hearings = load_data()
                result = pull_congress_api(hearings, api_key, silent=True)
            _last_pull["ts"] = datetime.now(timezone.utc).isoformat()
            _last_pull["new_api"] = len(result.get("new", []))
            _last_pull["api_updated"] = result.get("updated", 0)
            _last_pull["error"] = None
            new_n, upd_n = _last_pull["new_api"], _last_pull["api_updated"]
            if new_n and upd_n:
                _last_pull["message"] = f"Congress.gov: {new_n} new, {upd_n} updated."
            elif new_n:
                _last_pull["message"] = f"{new_n} new from Congress.gov API."
            elif upd_n:
                _last_pull["message"] = f"{upd_n} updated from Congress.gov API."
            else:
                _last_pull["message"] = "Congress.gov: no new or updated hearings."
        except Exception as e:
            _last_pull["error"] = str(e)
            _last_pull["message"] = str(e)
        finally:
            _last_pull["in_progress"] = False

    _last_pull["in_progress"] = True
    threading.Thread(target=_api_worker, name="CongressApiPull", daemon=True).start()
    return _flash_pull_started(url_for("feeds_page"))


@app.route("/feeds/api-key", methods=["POST"])
def save_api_key():
    # Clear when empty submit from Clear button
    if "api_key" in request.form and not request.form.get("api_key", "").strip():
        if os.path.exists(_API_KEY_FILE):
            os.remove(_API_KEY_FILE)
        flash("Congress.gov API key cleared.", "info")
        return redirect(url_for("feeds_page"))
    key = request.form.get("api_key", "").strip()
    if not key:
        flash("Enter a key to save, or use Clear to remove the stored key.", "error")
        return redirect(url_for("feeds_page"))
    _save_api_key(key)
    flash("Congress.gov API key saved. It is not shown again in the browser.", "success")
    return redirect(url_for("feeds_page"))


@app.route("/feeds/toggle/<int:idx>", methods=["POST"])
def toggle_feed(idx):
    with DATA_LOCK:
        feeds = load_feeds()
        if 0 <= idx < len(feeds):
            feeds[idx]["active"] = not feeds[idx].get("active", True)
            save_feeds(feeds)
    return redirect(url_for("feeds_page"))


@app.route("/feeds/delete/<int:idx>", methods=["POST"])
def delete_feed(idx):
    removed_name = None
    with DATA_LOCK:
        feeds = load_feeds()
        if 0 <= idx < len(feeds):
            removed_name = feeds.pop(idx)["name"]
            save_feeds(feeds)
    if removed_name:
        flash(f"Feed '{removed_name}' removed.", "info")
    return redirect(url_for("feeds_page"))


@app.route("/feeds/add", methods=["POST"])
def add_feed():
    with DATA_LOCK:
        feeds = load_feeds()
        committee = request.form.get("committee", "Other")
        if committee == "Other":
            custom = request.form.get("committee_custom", "").strip()
            if custom:
                committee = custom
        feeds.append({
            "name":      request.form["name"],
            "url":       request.form["url"],
            "committee": committee,
            "active":    True,
        })
        save_feeds(feeds)
    flash(f"Feed '{request.form['name']}' added.", "success")
    return redirect(url_for("feeds_page"))


@app.route("/calendar")
def chamber_calendar():
    view = request.args.get("view", "week")
    cache = load_chamber_calendar()
    if view == "month":
        ym = request.args.get("month", "")
        year, month = None, None
        if ym and "-" in ym:
            parts = ym.split("-", 1)
            try:
                year, month = int(parts[0]), int(parts[1])
            except ValueError:
                pass
        cal = build_session_calendar_month(year=year, month=month, cache=cache)
    else:
        cal = build_session_calendar_week(
            week_start=request.args.get("week") or None,
            cache=cache,
        )
    return render_template(
        "calendar.html",
        cal=cal,
        view=cal["view"],
        cache_updated=cache.get("updated"),
        official_links={
            "senate_schedule": cache.get("senate_source")
            or "https://www.senate.gov/legislative/2026_schedule.htm",
            "house_schedule": cache.get("house_source")
            or "https://www.house.gov/legislative-activity",
            "senate_floor": "https://www.senate.gov/legislative/schedule/weekly.htm",
        },
    )


@app.route("/calendar/pull", methods=["POST"])
def pull_calendar():
    view = request.form.get("view", "week")
    week = request.form.get("week", "")
    month = request.form.get("month", "")
    if view == "month":
        redirect_to = request.form.get("next") or url_for(
            "chamber_calendar", view="month", month=month
        )
    else:
        redirect_to = request.form.get("next") or url_for(
            "chamber_calendar", view="week", week=week
        )
    if _last_pull.get("in_progress"):
        flash("Another refresh is running. Please wait.", "error")
        return redirect(redirect_to)

    def _cal_worker():
        try:
            with DATA_LOCK:
                pull_session_calendar(silent=True)
            cache = load_chamber_calendar()
            n_house = len(cache.get("house_days") or {})
            n_senate = len(cache.get("senate_recess") or [])
            _last_pull["ts"] = datetime.now(timezone.utc).isoformat()
            _last_pull["error"] = None
            _last_pull["message"] = (
                f"Session calendar updated — House: {n_house} days, "
                f"Senate: {n_senate} recess periods."
            )
        except Exception as e:
            _last_pull["error"] = str(e)
            _last_pull["message"] = str(e)
        finally:
            _last_pull["in_progress"] = False

    _last_pull["in_progress"] = True
    threading.Thread(target=_cal_worker, name="SessionCalPull", daemon=True).start()
    flash("Session calendar refresh started in the background (~30 seconds).", "info")
    return redirect(redirect_to)


@app.route("/committees")
def committees_ongoings():
    hearings = load_data()
    cache = load_govtrack_cache()
    sections = build_committee_ongoings(hearings, cache)
    upcoming_total = sum(len(s["upcoming"]) for s in sections)
    return render_template(
        "committees.html",
        sections=sections,
        cache_updated=cache.get("updated"),
        upcoming_total=upcoming_total,
        govtrack_about="https://www.govtrack.us/about-our-data",
    )


@app.route("/committees/pull", methods=["POST"])
def pull_committees():
    redirect_to = request.form.get("next") or url_for("committees_ongoings")
    if not _start_background_feed_pull():
        flash("A refresh is already running. Please wait.", "error")
        return redirect(redirect_to)

    def _committee_meta_worker():
        try:
            with DATA_LOCK:
                refresh_govtrack_committee_cache(silent=True)
        except Exception as e:
            print(f"[govtrack-cache] {e}", file=sys.stderr, flush=True)

    threading.Thread(
        target=_committee_meta_worker, name="GovTrackCache", daemon=True
    ).start()
    return _flash_pull_started(redirect_to)


@app.route("/activity")
def activity():
    items  = load_social_items()
    feeds  = load_social_feeds()
    q      = request.args.get("q", "").lower()
    cf     = request.args.get("committee", "")
    if q:
        items = [i for i in items if q in i.get("title","").lower()
                 or q in i.get("summary","").lower()]
    if cf:
        items = [i for i in items if i.get("committee") == cf]
    items.sort(key=lambda x: x.get("date",""), reverse=True)
    return render_template(
        "activity.html",
        items=items,
        feeds=feeds,
        committees=COMMITTEES,
        q=request.args.get("q",""),
        cf=cf,
    )


@app.route("/activity/pull", methods=["POST"])
def pull_activity():
    if _last_pull.get("in_progress"):
        flash("A feed refresh is already running. Please wait.", "error")
        return redirect(url_for("activity"))
    with DATA_LOCK:
        feeds     = load_social_feeds()
        new_items = pull_social_feeds(feeds, silent=True)
    flash(f"{len(new_items)} new post(s) imported from social feeds.", "success")
    return redirect(url_for("activity"))


@app.route("/activity/feeds")
def activity_feeds():
    return render_template(
        "activity_feeds.html",
        feeds=load_social_feeds(),
        committees=COMMITTEES + ["Multiple"],
    )


@app.route("/activity/feeds/toggle/<int:idx>", methods=["POST"])
def toggle_social_feed(idx):
    with DATA_LOCK:
        feeds = load_social_feeds()
        if 0 <= idx < len(feeds):
            feeds[idx]["active"] = not feeds[idx].get("active", True)
            save_social_feeds(feeds)
    return redirect(url_for("activity_feeds"))


@app.route("/activity/feeds/delete/<int:idx>", methods=["POST"])
def delete_social_feed(idx):
    removed_name = None
    with DATA_LOCK:
        feeds = load_social_feeds()
        if 0 <= idx < len(feeds):
            removed_name = feeds.pop(idx)["name"]
            save_social_feeds(feeds)
    if removed_name:
        flash(f"Feed '{removed_name}' removed.", "info")
    return redirect(url_for("activity_feeds"))


@app.route("/activity/feeds/add", methods=["POST"])
def add_social_feed():
    with DATA_LOCK:
        feeds = load_social_feeds()
        committee = request.form.get("committee", "Other")
        if committee == "Other":
            custom = request.form.get("committee_custom", "").strip()
            if custom:
                committee = custom
        feeds.append({
            "name":      request.form["name"],
            "url":       request.form["url"],
            "committee": committee,
            "active":    True,
        })
        save_social_feeds(feeds)
    flash(f"Social feed '{request.form['name']}' added.", "success")
    return redirect(url_for("activity_feeds"))


@app.route("/activity/feeds/set-url/<int:idx>", methods=["POST"])
def set_social_feed_url(idx):
    msg = None
    with DATA_LOCK:
        feeds = load_social_feeds()
        if 0 <= idx < len(feeds):
            url = request.form.get("url", "").strip()
            feeds[idx]["url"]    = url
            feeds[idx]["active"] = bool(url)
            save_social_feeds(feeds)
            name = feeds[idx]["name"]
            msg = f"{'URL saved for' if url else 'URL cleared for'} {name}."
    if msg:
        flash(msg, "success")
    return redirect(url_for("activity_feeds"))


@app.route("/map")
def hearing_map():
    hearings = load_data()

    # Optional filters
    sf = request.args.get("status", "")
    cf = request.args.get("committee", "")
    filtered = hearings
    if sf:
        filtered = [h for h in filtered if h.get("status") == sf]
    if cf:
        filtered = [h for h in filtered if h.get("committee") == cf]

    points   = build_heatmap_points(filtered)
    buildings = {k: {"lat": v[0], "lng": v[1]} for k, v in DC_BUILDINGS.items()}

    return render_template(
        "map.html",
        points_json=_json.dumps(points),
        buildings_json=_json.dumps(buildings),
        total=len(points),
        committees=COMMITTEES,
        statuses=STATUS_OPTIONS,
        sf=sf, cf=cf,
    )


@app.route("/export")
def export():
    hearings = load_data()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("Run: pip install openpyxl", "error")
        return redirect(url_for("dashboard"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Hearing Tracker"

    headers = ["ID", "Date", "Days Away", "Status", "Source", "Committee",
               "Topic", "Witnesses", "Angle", "Action", "Questions", "Link", "Notes"]
    hdr_fill = PatternFill("solid", start_color="1F3864")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill("solid", start_color="DCE6F1")
    rss_fill = PatternFill("solid", start_color="E2EFDA")
    for row_num, h in enumerate(sorted(hearings, key=lambda x: x["date"]), 2):
        try:
            delta = (date.fromisoformat(h["date"]) - date.today()).days
            da    = "Today" if delta == 0 else (f"In {delta}d" if delta > 0 else f"{abs(delta)}d ago")
        except Exception:
            da = ""
        src_label = {"rss": "RSS", "api": "API"}.get(h.get("source", ""), "Manual")
        is_rss = h.get("source") in ("rss", "api")
        row = [h["id"], h["date"], da, h.get("status",""),
               src_label,
               h["committee"], h["topic"], h.get("witnesses",""),
               h.get("angle",""), h.get("action",""),
               h.get("questions","").replace(" | ", "\n"),
               h.get("url",""), h.get("notes","")]
        fill = rss_fill if is_rss else (alt_fill if row_num % 2 == 0 else None)
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    for col, w in enumerate([5,12,10,12,8,30,42,28,22,18,40,35,35], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"jamestown_hearings_{date.today()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


PORT = 5001

if __name__ == "__main__":
    is_bundled = getattr(sys, "frozen", False)
    print("\n  ┌──────────────────────────────────────────┐")
    print("  │  Jamestown Hearing Tracker — Web UI      │")
    print(f"  │  Open: http://localhost:{PORT}             │")
    print("  └──────────────────────────────────────────┘\n")

    # Keep data files next to the executable (or script) when bundled
    os.chdir(BASE_DIR)

    debug_mode = not is_bundled
    # Avoid duplicate poller threads from Flask’s debug reloader (parent vs child).
    if (not debug_mode) or os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        _start_auto_feed_poller()
        if RSS_AUTO_PULL_ENABLED:
            print(
                f"  RSS auto-pull: first run after {min(RSS_AUTO_PULL_INITIAL_DELAY_SEC, RSS_POLL_INTERVAL_MIN * 60)}s, "
                f"then every {RSS_POLL_INTERVAL_MIN} min "
                f"(RSS_AUTO_PULL=0 to disable; RSS_PULL_INTERVAL_MINUTES / RSS_AUTO_PULL_INITIAL_DELAY_SEC to tune)\n"
            )

    # Auto-open browser after a short delay
    threading.Timer(1.2, lambda: webbrowser.open(f"http://localhost:{PORT}")).start()

    # debug=False when bundled so the reloader doesn't interfere
    app.run(debug=debug_mode, port=PORT, use_reloader=not is_bundled)


def _auto_fr_poll_loop():
    import time
    interval_sec = FR_PULL_INTERVAL_MIN * 60
    time.sleep(min(RSS_AUTO_PULL_INITIAL_DELAY_SEC, interval_sec))
    while True:
        time.sleep(interval_sec)
        if not FR_AUTO_PULL_ENABLED:
            continue
        try:
            with DATA_LOCK:
                pull_federal_register(load_fr_documents(), load_fr_watchlist(), silent=True)
            print("[FR auto-pull] completed", flush=True)
        except Exception as e:
            print(f"[FR auto-pull] {e}", file=sys.stderr, flush=True)


def _start_auto_fr_poller():
    if not FR_AUTO_PULL_ENABLED:
        return
    threading.Thread(target=_auto_fr_poll_loop, name="FRAutoPull", daemon=True).start()


# Gunicorn / Render: start RSS poller in the WSGI process (use --workers 1).
if os.environ.get("RENDER") and RSS_AUTO_PULL_ENABLED:
    _start_auto_feed_poller()
if os.environ.get("RENDER") and FR_AUTO_PULL_ENABLED:
    _start_auto_fr_poller()
