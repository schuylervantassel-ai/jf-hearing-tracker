#!/usr/bin/env python3
"""
Congressional Hearing Tracker - Jamestown Foundation
Tracks hearings across key national security committees,
with automatic RSS feed pulling.

Requirements:
    pip install feedparser openpyxl
"""

import calendar
import json
import os
import re
import ssl
import sys
from html import unescape
from datetime import datetime, date
from urllib.parse import urlencode
from urllib.request import urlopen, Request
from urllib.error import URLError
import time
import xml.etree.ElementTree as ET

# ── Optional feedparser (better date parsing) ─────────────────────────────────
try:
    import feedparser
    HAS_FEEDPARSER = True
except ImportError:
    HAS_FEEDPARSER = False

# ── Storage ───────────────────────────────────────────────────────────────────
# Set DATA_DIR (e.g. /var/data on Render) for persistent disk; defaults to this folder.
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_DATA_ROOT = os.environ.get("DATA_DIR", "").strip() or _SCRIPT_DIR

def _data_path(filename: str) -> str:
    return os.path.join(_DATA_ROOT, filename)

DATA_FILE    = _data_path("hearings.json")
CONFIG_FILE  = _data_path("rss_config.json")
GOVTRACK_CACHE_FILE = _data_path("govtrack_committees.json")
CALENDAR_CACHE_FILE = _data_path("chamber_calendar.json")
SOCIAL_FILE  = _data_path("social_feed.json")
SENATE_SCHEDULE_PAGE = "https://www.senate.gov/legislative/{year}_schedule.htm"
HOUSE_SESSION_PAGE = "https://www.house.gov/legislative-activity/{year}-{month:02d}-01"
SENATE_SCHEDULE_URL = "https://www.senate.gov/general/committee_schedules/hearings.xml"
GOVTRACK_API_BASE = "https://www.govtrack.us/api/v2"
SOCIAL_LIMIT = 500   # max items to keep across all social feeds

# ── RSS Feeds ─────────────────────────────────────────────────────────────────
DEFAULT_FEEDS = [
    {
        "name": "Senate Armed Services (SASC)",
        "url": "https://rss.app/feeds/TQ2HZ3EcaUZdjfZk.xml",
        "committee": "Senate Armed Services (SASC)",
        "active": True,
    },
    {
        "name": "Senate Foreign Relations (SFRC)",
        "url": "https://rss.app/feeds/52g7oTFDxPrMSp15.xml",
        "committee": "Senate Foreign Relations (SFRC)",
        "active": True,
    },
    {
        "name": "Senate Intelligence (SSCI)",
        "url": "https://rss.app/feeds/Wk9rFjawif2PF1cC.xml",
        "committee": "Senate Intelligence (SSCI)",
        "active": True,
    },
    {
        "name": "Senate Homeland Security (SHSGAC)",
        "url": "https://rss.app/feeds/viBWi7HQxZLlJ85m.xml",
        "committee": "Senate Homeland Security (SHSGAC)",
        "active": True,
    },
    {
        "name": "House Armed Services (HASC)",
        "url": "https://armedservices.house.gov/rss.xml",
        "committee": "House Armed Services (HASC)",
        "active": False,  # No public RSS feed available; add manually
    },
    {
        "name": "House Foreign Affairs (HFAC)",
        "url": "https://rss.app/feeds/iqYqfkbDWR3Q2Y4i.xml",
        "committee": "House Foreign Affairs (HFAC)",
        "active": True,
    },
    {
        "name": "House Intelligence (HPSCI)",
        "url": "https://intelligence.house.gov/feed/",
        "committee": "House Intelligence (HPSCI)",
        "active": True,
    },
    {
        "name": "House Homeland Security (HHSC)",
        "url": "https://homeland.house.gov/feed/",
        "committee": "House Homeland Security (HHSC)",
        "active": True,
    },
    {
        "name": "GovInfo - All Congressional Hearings",
        "url": "https://www.govinfo.gov/rss/chrg.xml",
        "committee": "Multiple",
        "active": True,
    },
]

# Social / X (Twitter) feeds — tracked separately from hearing RSS feeds
# url="" means not yet configured; set active=True once a url is added.
DEFAULT_SOCIAL_FEEDS = [
    {
        "name": "SASC on X",
        "handle": "@ArmedServices",
        "committee": "Senate Armed Services (SASC)",
        "url": "",
        "active": False,
    },
    {
        "name": "SFRC on X",
        "handle": "@SenateForeign",
        "committee": "Senate Foreign Relations (SFRC)",
        "url": "",
        "active": False,
    },
    {
        "name": "SSCI on X",
        "handle": "@SenIntelComm",
        "committee": "Senate Intelligence (SSCI)",
        "url": "",
        "active": False,
    },
    {
        "name": "SHSGAC on X",
        "handle": "@SenHomeland",
        "committee": "Senate Homeland Security (SHSGAC)",
        "url": "",
        "active": False,
    },
    {
        "name": "HASC on X",
        "handle": "@HASCRepublicans",
        "committee": "House Armed Services (HASC)",
        "url": "",
        "active": False,
    },
    {
        "name": "HFAC Majority on X",
        "handle": "@HouseForeignGOP",
        "committee": "House Foreign Affairs (HFAC)",
        "url": "https://rss.app/feeds/wJ2Wr7h4hpDCsOgi.xml",
        "active": True,
    },
    {
        "name": "HPSCI on X",
        "handle": "@HouseIntelGOP",
        "committee": "House Intelligence (HPSCI)",
        "url": "",
        "active": False,
    },
    {
        "name": "HHSC on X",
        "handle": "@HomelandGOP",
        "committee": "House Homeland Security (HHSC)",
        "url": "",
        "active": False,
    },
]

SOCIAL_CONFIG_FILE = _data_path("social_config.json")

COMMITTEES = [
    "Senate Armed Services (SASC)",
    "Senate Foreign Relations (SFRC)",
    "Senate Intelligence (SSCI)",
    "Senate Homeland Security (SHSGAC)",
    "House Armed Services (HASC)",
    "House Foreign Affairs (HFAC)",
    "House Intelligence (HPSCI)",
    "House Homeland Security (HHSC)",
    "Other",
]

JAMESTOWN_ANGLES = [
    "Russia/Eurasia",
    "China/Indo-Pacific",
    "Middle East/North Africa",
    "Sub-Saharan Africa",
    "Terrorism/Extremism",
    "Central Asia",
    "Latin America",
    "Cyber/Information Warfare",
    "Other",
]

ACTIONS = [
    "Send brief",
    "Send questions",
    "Offer testimony",
    "Request meeting",
    "Monitor only",
    "No action needed",
]

STATUS_OPTIONS = ["Upcoming", "Completed", "Cancelled", "Postponed"]

# Keywords that auto-flag a hearing's Jamestown angle
RELEVANCE_KEYWORDS = {
    "Russia/Eurasia": ["russia", "ukraine", "nato", "eurasia", "belarus", "moldova",
                       "caucasus", "georgia", "kremlin", "wagner", "putin"],
    "China/Indo-Pacific": ["china", "taiwan", "indo-pacific", "pla", "beijing",
                            "south china sea", "xinjiang", "tibet", "hong kong"],
    "Middle East/North Africa": ["iran", "iraq", "syria", "yemen", "israel", "hamas",
                                  "hezbollah", "mena", "saudi", "gulf", "libya", "egypt"],
    "Sub-Saharan Africa": ["africa", "sahel", "mali", "niger", "nigeria", "somalia",
                            "al-shabaab", "sudan", "ethiopia"],
    "Terrorism/Extremism": ["terrorism", "isis", "al-qaeda", "extremism", "jihadist",
                             "counterterrorism", "violent extremism", "radicalization"],
    "Central Asia": ["kazakhstan", "uzbekistan", "tajikistan", "kyrgyzstan",
                     "turkmenistan", "afghanistan", "central asia"],
    "Latin America": ["venezuela", "cuba", "nicaragua", "cartel", "narco",
                      "latin america", "mexico", "colombia"],
    "Cyber/Information Warfare": ["cyber", "disinformation", "information warfare",
                                   "influence operation", "hack", "ransomware", "espionage"],
}

# ── Data helpers ──────────────────────────────────────────────────────────────

def auto_close_past_hearings(hearings):
    """
    Set status to 'Completed' for any hearing that was 'Upcoming'
    but whose date is strictly before today. Modifies list in place
    and returns the number of records changed.
    """
    today = date.today().isoformat()
    changed = 0
    for h in hearings:
        if h.get("status") == "Upcoming" and h.get("date", "9999") < today:
            h["status"] = "Completed"
            changed += 1
    return changed

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE) as f:
            hearings = json.load(f)
        # Auto-close any hearings whose date has passed
        if auto_close_past_hearings(hearings):
            save_data(hearings)
        return hearings
    return []

def save_data(hearings):
    """Atomic write so readers never see a half-written hearings.json."""
    tmp = DATA_FILE + ".tmp"
    with open(tmp, "w") as f:
        json.dump(hearings, f, indent=2)
    os.replace(tmp, DATA_FILE)

def load_feeds():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE) as f:
            return json.load(f)
    save_feeds(DEFAULT_FEEDS)
    return DEFAULT_FEEDS

def save_feeds(feeds):
    with open(CONFIG_FILE, "w") as f:
        json.dump(feeds, f, indent=2)

def load_social_feeds():
    if not os.path.exists(SOCIAL_CONFIG_FILE):
        save_social_feeds(DEFAULT_SOCIAL_FEEDS)
        return [dict(f) for f in DEFAULT_SOCIAL_FEEDS]
    with open(SOCIAL_CONFIG_FILE) as f:
        saved = json.load(f)
    # Merge: add any DEFAULT entries whose committee isn't represented yet
    saved_committees = {f["committee"] for f in saved}
    changed = False
    for default in DEFAULT_SOCIAL_FEEDS:
        if default["committee"] not in saved_committees:
            saved.append(dict(default))
            changed = True
    if changed:
        save_social_feeds(saved)
    return saved

def save_social_feeds(feeds):
    with open(SOCIAL_CONFIG_FILE, "w") as f:
        json.dump(feeds, f, indent=2)

def load_social_items():
    if os.path.exists(SOCIAL_FILE):
        with open(SOCIAL_FILE) as f:
            return json.load(f)
    return []

def save_social_items(items):
    with open(SOCIAL_FILE, "w") as f:
        json.dump(items, f, indent=2)

def next_id(hearings):
    return max((h["id"] for h in hearings), default=0) + 1

# ── Display helpers ───────────────────────────────────────────────────────────

def divider(char="─", width=72):
    print(char * width)

def header(title):
    print()
    divider("═")
    print(f"  {title}")
    divider("═")

def pick(prompt, options, allow_blank=False):
    print(f"\n{prompt}")
    for i, opt in enumerate(options, 1):
        print(f"  {i}. {opt}")
    if allow_blank:
        print("  [Enter to skip]")
    while True:
        raw = input("  Choice: ").strip()
        if allow_blank and raw == "":
            return None
        if raw.isdigit() and 1 <= int(raw) <= len(options):
            return options[int(raw) - 1]
        print("  Invalid choice, try again.")

def ask(prompt, required=True, default=None):
    hint = f" [{default}]" if default else ""
    while True:
        val = input(f"  {prompt}{hint}: ").strip()
        if val:
            return val
        if default:
            return default
        if not required:
            return ""
        print("  This field is required.")

def ask_date(prompt, required=True):
    while True:
        raw = input(f"  {prompt} (YYYY-MM-DD): ").strip()
        if not raw and not required:
            return None
        try:
            datetime.strptime(raw, "%Y-%m-%d")
            return raw
        except ValueError:
            print("  Invalid date. Use YYYY-MM-DD.")

def days_away_str(date_str):
    try:
        delta = (date.fromisoformat(date_str) - date.today()).days
        if delta == 0:
            return " <- TODAY"
        elif delta > 0:
            return f" (in {delta}d)"
        else:
            return f" ({abs(delta)}d ago)"
    except Exception:
        return ""

def format_hearing(h, verbose=False):
    source_tag = " [RSS]" if h.get("source") == "rss" else " [Manual]"
    print(f"\n  ID {h['id']:03d} | {h['date']}{days_away_str(h['date'])} | [{h.get('status', 'Upcoming')}]{source_tag}")
    print(f"  Committee : {h['committee']}")
    print(f"  Topic     : {h['topic']}")
    if h.get("witnesses"):
        print(f"  Witnesses : {h['witnesses']}")
    if h.get("url"):
        print(f"  Link      : {h['url']}")
    print(f"  Angle     : {h.get('angle', '--')}")
    print(f"  Action    : {h.get('action', '--')}")
    if verbose:
        if h.get("questions"):
            print(f"\n  Questions to ask:")
            for q in h["questions"].split("|"):
                q = q.strip()
                if q:
                    print(f"    * {q}")
        if h.get("notes"):
            print(f"\n  Notes     : {h['notes']}")
    divider()

# ── RSS Parsing ───────────────────────────────────────────────────────────────

def detect_angle(text):
    text_lower = text.lower()
    for angle, keywords in RELEVANCE_KEYWORDS.items():
        if any(kw in text_lower for kw in keywords):
            return angle
    return "Other"

def parse_date_str(date_str):
    if not date_str:
        return date.today().isoformat()
    formats = [
        "%a, %d %b %Y %H:%M:%S %z",
        "%a, %d %b %Y %H:%M:%S GMT",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%d",
        "%B %d, %Y",
        "%b %d, %Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    if HAS_FEEDPARSER:
        parsed = feedparser._parse_date(date_str)
        if parsed:
            return date(*parsed[:3]).isoformat()
    return date.today().isoformat()

def clean_html(text):
    return re.sub(r"<[^>]+>", "", text or "").strip()

def _congress_year(n):
    """Return the start year of the Nth Congress (119th → 2025)."""
    return 2025 - 2 * (119 - n)

def _extract_congress_date(title, url):
    """
    Some feeds (e.g. GovInfo) assign today's pubDate to old records they've
    just digitised.  Detect the actual hearing era from the Congress number
    embedded in the title or URL and return the END date of that Congress's
    term — so the filter only rejects items where the entire Congress is
    older than the cutoff.

    Examples:
      CHRG-119... → 119th Congress (2025-2027, ongoing) → today's date → KEEP
      CHRG-118... → 118th Congress ended Jan 2025 → '2025-01-03' → SKIP if >90d
      CHRG-117... → 117th Congress ended Jan 2023 → '2023-01-03' → SKIP

    Returns None if no Congress number is detected.
    """
    combined = f"{title or ''} {url or ''}"
    for pat in (
        r"CHRG-(\d{2,3})[a-z]",              # CHRG-119hhrg… / CHRG-117shrg…
        r"[Hh]rg[.\s]+(\d{3})\s*[-\u2013]",  # S. Hrg. 119-… / Hrg. 117-…
        r"(\d{3})(?:st|nd|rd|th)\s+[Cc]ongress",  # 119th Congress
    ):
        m = re.search(pat, combined)
        if m:
            n = int(m.group(1))
            if 80 <= n <= 130:  # sanity check
                end_date = date(_congress_year(n) + 2, 1, 3)
                # If the Congress is still ongoing, use today so it always passes
                if end_date > date.today():
                    end_date = date.today()
                return end_date.isoformat()
    return None

def is_duplicate(hearings, title, committee):
    title_lower = title.lower().strip()
    for h in hearings:
        if h["committee"] == committee and h["topic"].lower().strip() == title_lower:
            return True
    return False

def _make_ssl_context():
    try:
        import certifi
        return ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        pass
    ctx = ssl.create_default_context()
    try:
        ctx.load_default_certs()
    except Exception:
        pass
    return ctx

def fetch_feed_raw(url):
    headers = {"User-Agent": "Mozilla/5.0 (Jamestown Foundation Hearing Tracker)"}
    req = Request(url, headers=headers)
    ctx = _make_ssl_context()
    try:
        with urlopen(req, timeout=10, context=ctx) as resp:
            content = resp.read()
    except ssl.SSLCertVerificationError:
        fallback = ssl.create_default_context()
        fallback.check_hostname = False
        fallback.verify_mode = ssl.CERT_NONE
        with urlopen(req, timeout=10, context=fallback) as resp:
            content = resp.read()

    if HAS_FEEDPARSER:
        feed = feedparser.parse(content)
        items = []
        for entry in feed.entries:
            items.append({
                "title":   entry.get("title", ""),
                "date":    entry.get("published", entry.get("updated", "")),
                "url":     entry.get("link", ""),
                "summary": entry.get("summary", entry.get("description", "")),
            })
        return items

    # Fallback: raw XML
    root = ET.fromstring(content)
    ns   = {"atom": "http://www.w3.org/2005/Atom"}
    items = []
    for item in root.findall(".//item"):
        items.append({
            "title":   item.findtext("title", ""),
            "date":    item.findtext("pubDate", ""),
            "url":     item.findtext("link", ""),
            "summary": item.findtext("description", ""),
        })
    if not items:
        for entry in root.findall(".//atom:entry", ns):
            link_el = entry.find("atom:link", ns)
            items.append({
                "title":   entry.findtext("atom:title", "", ns),
                "date":    entry.findtext("atom:updated", "", ns),
                "url":     link_el.get("href", "") if link_el is not None else "",
                "summary": entry.findtext("atom:summary", "", ns),
            })
    return items

RSS_CUTOFF_DAYS = 90   # ignore feed items older than this

def pull_rss_feeds(hearings, feeds, silent=False, persist=True):
    new_items = []
    cutoff = date.today().toordinal() - RSS_CUTOFF_DAYS
    seen_topics = {
        (h.get("committee"), (h.get("topic") or "").lower().strip())
        for h in hearings
    }

    for feed in feeds:
        if not feed.get("active", True):
            continue
        if not silent:
            print(f"  Fetching: {feed['name']} ... ", end="", flush=True)
        try:
            items = fetch_feed_raw(feed["url"])
            added = 0
            skipped_old = 0
            for item in items:
                title   = clean_html(item["title"])
                summary = clean_html(item["summary"])
                if not title:
                    continue
                # Drop items older than the cutoff (by pubDate)
                item_date_str = parse_date_str(item["date"])
                try:
                    item_ord = date.fromisoformat(item_date_str).toordinal()
                except Exception:
                    item_ord = cutoff
                if item_ord < cutoff:
                    skipped_old += 1
                    continue

                # Secondary check: detect actual hearing era from Congress
                # number in URL/title (GovInfo backfills old records with
                # current pubDates, so pubDate alone is not enough).
                content_date = _extract_congress_date(title, item.get("url", ""))
                if content_date:
                    try:
                        if date.fromisoformat(content_date).toordinal() < cutoff:
                            skipped_old += 1
                            continue
                    except Exception:
                        pass

                dup_key = (feed["committee"], title.lower().strip())
                if dup_key in seen_topics:
                    continue
                seen_topics.add(dup_key)
                angle = detect_angle(f"{title} {summary}")
                item_date_resolved = parse_date_str(item["date"])
                auto_status = "Completed" if item_date_resolved < date.today().isoformat() else "Upcoming"
                h = {
                    "id":        next_id(hearings + new_items),
                    "date":      item_date_resolved,
                    "committee": feed["committee"],
                    "topic":     title,
                    "witnesses": "",
                    "angle":     angle,
                    "action":    "Monitor only",
                    "status":    auto_status,
                    "questions": "",
                    "notes":     summary[:300] if summary else "",
                    "url":       normalize_external_url(item.get("url", "")),
                    "source":    "rss",
                    "created":   datetime.now().strftime("%Y-%m-%d"),
                }
                new_items.append(h)
                hearings.append(h)
                added += 1
            if not silent:
                old_note = f", {skipped_old} too old" if skipped_old else ""
                print(f"{added} new item(s){old_note}")
            feed["last_fetched"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            feed["last_status"]  = "OK"
        except URLError as e:
            if not silent:
                print(f"FAILED (network: {e.reason})")
            feed["last_status"] = f"Error: {e.reason}"
        except Exception as e:
            if not silent:
                print(f"FAILED ({e})")
            feed["last_status"] = f"Error: {e}"

    if persist:
        save_data(hearings)
        save_feeds(feeds)
    return new_items

# ── GovTrack + Senate schedule (committee on-goings) ─────────────────────────

# Canonical HTTPS committee home pages (overrides stale GovTrack http:// URLs).
OFFICIAL_COMMITTEE_SITES = {
    "Senate Armed Services (SASC)": "https://www.armed-services.senate.gov/",
    "Senate Foreign Relations (SFRC)": "https://www.foreign.senate.gov/",
    "Senate Intelligence (SSCI)": "https://www.intelligence.senate.gov/",
    "Senate Homeland Security (SHSGAC)": "https://www.hsgac.senate.gov/",
    "House Armed Services (HASC)": "https://armedservices.house.gov/",
    "House Foreign Affairs (HFAC)": "https://foreignaffairs.house.gov/",
    "House Intelligence (HPSCI)": "https://intelligence.house.gov/",
    "House Homeland Security (HHSC)": "https://homeland.house.gov/",
}

_SENATE_ISVP_RE = re.compile(r"https?://www\.senate\.gov/isvp/[^\s;]+", re.I)


def _committee_site_urls(cfg=None):
    urls = dict(OFFICIAL_COMMITTEE_SITES)
    if cfg is None:
        cfg = load_congress_config()
    for entry in cfg.get("committees") or []:
        label = entry.get("label", "")
        site = (entry.get("site_url") or "").strip()
        if label and site:
            urls[label] = site
    return urls


def normalize_external_url(url):
    """Force https and decode XML entities in outbound links."""
    if not url:
        return ""
    url = unescape(str(url).strip())
    if url.startswith("http://"):
        url = "https://" + url[7:]
    return url


def _senate_schedule_meeting_url(video_url, congress, event_id):
    """Prefer Senate ISVP stream/page from schedule XML over Congress.gov."""
    video = normalize_external_url(video_url)
    if video and "senate.gov" in video:
        return video
    if event_id:
        return (
            f"https://www.congress.gov/committee-meeting/"
            f"{congress}/senate/{event_id}"
        )
    return ""


_LEGACY_CONGRESS_MTG_RE = re.compile(
    r"^https://www\.congress\.gov/committee-meeting/(\d+)/(house|senate)/(\d+)$",
    re.I,
)


def _chamber_slug(chamber):
    c = (chamber or "").lower()
    return "house" if c in ("house", "h") else "senate"


def _congress_event_page_url(congress, chamber, event_id):
    """Public Congress.gov event page (API uses this form in meeting videos)."""
    ch = _chamber_slug(chamber)
    return f"https://www.congress.gov/event/{congress}th-Congress/{ch}-event/{event_id}"


def _urls_from_api_videos(meeting):
    """Prefer Congress.gov event pages from API videos; ISVP is a last resort."""
    congress_event = ""
    senate_stream = ""
    fallback = ""
    for v in (meeting or {}).get("videos") or []:
        u = normalize_external_url(v.get("url") or "")
        if not u or "api.congress.gov" in u:
            continue
        if "congress.gov/event" in u:
            congress_event = u
        elif "senate.gov/isvp" in u or "senate.gov" in u:
            senate_stream = u
        elif not fallback and "congress.gov" not in u:
            fallback = u
    return congress_event or senate_stream or fallback


def _congress_api_meeting_url(meeting, congress, chamber, event_id):
    """
    Congress.gov event page for API imports (hearing details, documents).
    Senate ISVP stream links are only used when no event page is available.
    """
    congress_event = ""
    senate_stream = ""
    for v in (meeting or {}).get("videos") or []:
        u = normalize_external_url(v.get("url") or "")
        if not u or "api.congress.gov" in u:
            continue
        if "congress.gov/event" in u:
            congress_event = u
        elif "senate.gov/isvp" in u or "senate.gov" in u:
            senate_stream = u
    if congress_event:
        return congress_event
    if event_id:
        return _congress_event_page_url(congress, chamber, event_id)
    for key in ("url", "meetingUrl", "webLink"):
        raw = normalize_external_url((meeting or {}).get(key) or "")
        if not raw or "api.congress.gov" in raw:
            continue
        if "congress.gov/event" in raw:
            return raw
    if senate_stream:
        return senate_stream
    for key in ("url", "meetingUrl", "webLink"):
        raw = normalize_external_url((meeting or {}).get(key) or "")
        if not raw or "api.congress.gov" in raw:
            continue
        if any(host in raw for host in ("congress.gov", "house.gov", "senate.gov")):
            return raw
    return ""


def repair_stored_hearing_urls(hearings):
    """
    Fix URLs already on disk: schedule rows → Senate ISVP when in notes,
    and upgrade http:// links to https://.
    """
    changed = 0
    for h in hearings:
        old = h.get("url") or ""
        new = normalize_external_url(old)
        if h.get("source") == "schedule":
            m = _SENATE_ISVP_RE.search(h.get("notes") or "")
            if m:
                new = normalize_external_url(m.group(0))
        elif h.get("source") == "api":
            m = _LEGACY_CONGRESS_MTG_RE.match(new or old)
            if m:
                new = _congress_event_page_url(m.group(1), m.group(2), m.group(3))
            elif h.get("congress_event_id") and "senate.gov/isvp" in (new or old):
                cfg = load_congress_config()
                ch = _chamber_slug(
                    "senate" if "Senate" in (h.get("committee") or "") else "house"
                )
                new = _congress_event_page_url(
                    cfg["congress"], ch, str(h["congress_event_id"])
                )
        if new and new != old:
            h["url"] = new
            changed += 1
    if changed:
        save_data(hearings)
    return changed


def fetch_govtrack_json(path):
    """GET GovTrack API v2 (no key required). path e.g. '/committee?code=SSAS'."""
    url = GOVTRACK_API_BASE + path
    headers = {"User-Agent": "Mozilla/5.0 (Jamestown Foundation Hearing Tracker)"}
    req = Request(url, headers=headers)
    ctx = _make_ssl_context()
    try:
        with urlopen(req, timeout=12, context=ctx) as resp:
            return json.loads(resp.read())
    except ssl.SSLCertVerificationError:
        fallback = ssl.create_default_context()
        fallback.check_hostname = False
        fallback.verify_mode = ssl.CERT_NONE
        with urlopen(req, timeout=12, context=fallback) as resp:
            return json.loads(resp.read())


def load_govtrack_cache():
    if os.path.exists(GOVTRACK_CACHE_FILE):
        with open(GOVTRACK_CACHE_FILE) as f:
            return json.load(f)
    return {}


def save_govtrack_cache(data):
    with open(GOVTRACK_CACHE_FILE, "w") as f:
        json.dump(data, f, indent=2)


def refresh_govtrack_committee_cache(silent=False):
    """Pull committee metadata from GovTrack.us for each tracked committee."""
    cfg = load_congress_config()
    prior = load_govtrack_cache()
    cache = {
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "committees": dict(prior.get("committees") or {}),
    }
    site_urls = _committee_site_urls(cfg)
    for entry in cfg.get("committees") or []:
        code = entry.get("govtrack_code") or ""
        label = entry.get("label", "")
        if not code:
            continue
        if not silent:
            print(f"  GovTrack ({label}) ... ", end="", flush=True)
        try:
            data = fetch_govtrack_json(f"/committee?code={code}&limit=1")
            objs = data.get("objects") or []
            if not objs:
                if not silent:
                    print("not found")
                continue
            o = objs[0]
            cache["committees"][label] = {
                "govtrack_code": code,
                "name": o.get("name", label),
                "abbrev": o.get("abbrev", ""),
                "jurisdiction": o.get("jurisdiction", ""),
                "jurisdiction_link": o.get("jurisdiction_link", ""),
                "official_url": site_urls.get(label) or normalize_external_url(o.get("url", "")),
                "govtrack_url": f"https://www.govtrack.us/congress/committees/{code}",
                "chamber": o.get("committee_type", entry.get("chamber", "")),
            }
            if not silent:
                print("OK")
        except Exception as e:
            if not silent:
                print(f"FAILED ({e})")
    save_govtrack_cache(cache)
    return cache


def _fetch_senate_schedule_xml():
    headers = {"User-Agent": "Mozilla/5.0 (Jamestown Foundation Hearing Tracker)"}
    url = load_congress_config().get("senate_schedule_url") or SENATE_SCHEDULE_URL
    req = Request(url, headers=headers)
    ctx = _make_ssl_context()
    with urlopen(req, timeout=15, context=ctx) as resp:
        return resp.read()


def pull_senate_schedule(hearings, silent=False, persist=True):
    """
    Import upcoming Senate committee meetings from the official schedule XML
    (same source GovTrack uses — see govtrack.us/about-our-data).
    """
    cfg = load_congress_config()
    code_map = cfg["committee_codes"]
    congress = cfg["congress"]
    new_items = []
    updated = 0
    if not silent:
        print("  Senate committee schedule (XML) ... ", end="", flush=True)
    try:
        root = ET.fromstring(_fetch_senate_schedule_xml())
    except Exception as e:
        if not silent:
            print(f"FAILED ({e})")
        return {"new": [], "updated": 0}

    today = date.today().isoformat()
    for mtg in root.findall("meeting"):
        cmte_code = (mtg.findtext("cmte_code") or "").lower()
        committee_name = _committee_from_system_code(cmte_code, code_map)
        if not committee_name:
            continue
        event_id = (mtg.findtext("identifier") or "").strip()
        matter = clean_html(mtg.findtext("matter") or "")
        if not matter and not event_id:
            continue
        meeting_date = (mtg.findtext("date_iso_8601") or "")[:10]
        if not meeting_date or not meeting_date[0].isdigit():
            meeting_date = date.today().isoformat()
        title = matter[:500] if matter else f"{committee_name} – Senate meeting"
        sub = mtg.findtext("sub_cmte") or ""
        room = mtg.findtext("room") or ""
        video = mtg.findtext("video_url") or ""
        notes_parts = []
        if sub:
            notes_parts.append(sub)
        if room:
            notes_parts.append(room)
        if video:
            notes_parts.append(video)
        notes = "; ".join(notes_parts)

        status = "Upcoming" if meeting_date >= today else "Completed"
        hearing = _find_hearing_for_api(hearings, event_id, committee_name, title)
        if hearing:
            hearing["committee"] = committee_name
            hearing["topic"] = title
            hearing["date"] = meeting_date
            hearing["status"] = status
            hearing["notes"] = notes
            hearing["url"] = _senate_schedule_meeting_url(video, congress, event_id)
            hearing["congress_event_id"] = event_id
            hearing["source"] = "schedule"
            hearing["angle"] = detect_angle(title)
            updated += 1
        else:
            if is_duplicate(hearings + new_items, title, committee_name):
                continue
            h = {
                "id": next_id(hearings + new_items),
                "date": meeting_date,
                "committee": committee_name,
                "topic": title,
                "witnesses": "",
                "angle": detect_angle(title),
                "action": "Monitor only",
                "status": status,
                "questions": "",
                "notes": notes,
                "url": _senate_schedule_meeting_url(video, congress, event_id),
                "congress_event_id": event_id,
                "source": "schedule",
                "created": date.today().isoformat(),
            }
            new_items.append(h)
            hearings.append(h)

    if persist:
        repair_stored_hearing_urls(hearings)
        save_data(hearings)
    if not silent:
        print(f"{len(new_items)} new, {updated} updated")
    return {"new": new_items, "updated": updated}


# ── House / Senate session calendar (in session vs recess) ───────────────────

_MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _fetch_url_bytes(url, timeout=15):
    headers = {"User-Agent": "Mozilla/5.0 (Jamestown Foundation Hearing Tracker)"}
    req = Request(url, headers=headers)
    ctx = _make_ssl_context()
    with urlopen(req, timeout=timeout, context=ctx) as resp:
        return resp.read()


def _parse_month_day_token(token, year):
    token = re.sub(r"&nbsp;", " ", (token or "")).strip()
    m = re.match(r"([A-Za-z]+)\s*(\d{1,2})", token)
    if not m:
        return None
    mo = _MONTH_MAP.get(m.group(1)[:3].lower())
    if not mo:
        return None
    try:
        return date(year, mo, int(m.group(2)))
    except ValueError:
        return None


def _parse_senate_recess_periods(year):
    """Tentative recess / state work periods from Senate.gov schedule page."""
    url = SENATE_SCHEDULE_PAGE.format(year=year)
    html = _fetch_url_bytes(url).decode("utf-8", errors="replace")
    periods = []
    for row in re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S | re.I):
        cells = re.findall(r"<td[^>]*>(.*?)</td>", row, re.S | re.I)
        if len(cells) < 2:
            continue
        datecell = re.sub(r"<[^>]+>", " ", cells[0])
        datecell = re.sub(r"\s+", " ", datecell).strip()
        action = re.sub(r"<[^>]+>", " ", cells[1])
        action = re.sub(r"\s+", " ", action).strip()
        note = ""
        if len(cells) > 2:
            note = re.sub(r"<[^>]+>", " ", cells[2])
            note = re.sub(r"\s+", " ", note).strip()
        if not re.search(r"\d", datecell) or datecell.lower().startswith("date"):
            continue
        m = re.match(r"(.+?)\s*-\s*(.+)", datecell)
        if not m:
            continue
        d1 = _parse_month_day_token(m.group(1), year)
        d2 = _parse_month_day_token(m.group(2), year)
        if not d1 or not d2:
            continue
        label = action or note or "Recess"
        periods.append({
            "start": d1.isoformat(),
            "end": d2.isoformat(),
            "label": label,
        })
    return periods, url


def _house_class_to_status(css_class):
    css = (css_class or "").strip().lower()
    if "in-session" in css:
        return "in_session", "In session"
    if "district-work" in css:
        return "recess", "District work period"
    if "holiday" in css or "federal" in css:
        return "holiday", "Federal holiday"
    if not css:
        return "recess", "Not in session"
    return "unknown", css.replace("-", " ").title()


def _parse_house_session_month(html, year, month):
    """Extract per-day House status from house.gov legislative-activity calendar."""
    days = {}
    for table in re.findall(
        r'<table class="housegov-calendar"[^>]*id="housegov-calendar-'
        r'(\d+)-(\d+)"[^>]*>(.*?)</table>',
        html,
        re.S | re.I,
    ):
        tbl_month, tbl_year = int(table[0]), int(table[1])
        if tbl_year != year or tbl_month != month:
            continue
        for td in re.findall(
            r'<td class="([^"]*)"[^>]*>(.*?)</td>',
            table[2],
            re.S | re.I,
        ):
            m_date = re.search(
                r'class="screen-reader-text">([A-Za-z]+ \d{1,2}, \d{4})',
                td[1],
            )
            if not m_date:
                continue
            try:
                d = datetime.strptime(m_date.group(1), "%B %d, %Y").date()
            except ValueError:
                continue
            status, label = _house_class_to_status(td[0])
            days[d.isoformat()] = {"status": status, "label": label}
    return days


def _parse_house_session_calendar(year, silent=False):
    """Fetch House.gov calendar tables (main page + monthly URLs)."""
    all_days = {}
    errors = []
    try:
        html_main = _fetch_url_bytes(
            "https://www.house.gov/legislative-activity", timeout=20
        ).decode("utf-8", errors="replace")
        for month in range(1, 13):
            all_days.update(_parse_house_session_month(html_main, year, month))
    except Exception as e:
        errors.append(f"main: {e}")

    # Monthly URLs only needed if main page lacks that month (House shows ~3 months).
    months_missing = [m for m in range(1, 13) if not any(
        d.startswith(f"{year}-{m:02d}") for d in all_days
    )]
    for month in months_missing:
        url = HOUSE_SESSION_PAGE.format(year=year, month=month)
        if not silent:
            print(f"    House {year}-{month:02d} ... ", end="", flush=True)
        try:
            html = _fetch_url_bytes(url, timeout=20).decode("utf-8", errors="replace")
            days = _parse_house_session_month(html, year, month)
            all_days.update(days)
            if not silent:
                print(f"{len(days)} days")
            time.sleep(0.15)
        except Exception as e:
            errors.append(f"{month:02d}: {e}")
            if not silent:
                print(f"FAILED ({e})")
    return all_days, errors


def _senate_status_for_day(d, recess_periods):
    iso = d.isoformat()
    for p in recess_periods:
        if p["start"] <= iso <= p["end"]:
            return "recess", p["label"]
    if d.weekday() >= 5:
        return "recess", "Weekend"
    return "in_session", "Expected in session"


def load_chamber_calendar():
    if os.path.exists(CALENDAR_CACHE_FILE):
        with open(CALENDAR_CACHE_FILE) as f:
            return json.load(f)
    return {"updated": None, "year": date.today().year, "senate_recess": [], "house_days": {}}


def save_chamber_calendar(data):
    with open(CALENDAR_CACHE_FILE, "w") as f:
        json.dump(data, f, indent=2)


def pull_session_calendar(silent=False, year=None):
    """Refresh in-session / recess data for House and Senate."""
    year = year or date.today().year
    errors = []

    if not silent:
        print(f"  Senate {year} tentative schedule ... ", end="", flush=True)
    try:
        senate_recess, senate_url = _parse_senate_recess_periods(year)
        if not silent:
            print(f"{len(senate_recess)} recess period(s)")
    except Exception as e:
        senate_recess, senate_url = [], SENATE_SCHEDULE_PAGE.format(year=year)
        errors.append(f"Senate: {e}")
        if not silent:
            print(f"FAILED ({e})")

    if not silent:
        print(f"  House {year} session calendar ...")
    try:
        house_days, house_errors = _parse_house_session_calendar(year, silent=silent)
        errors.extend(house_errors[:5])
    except Exception as e:
        house_days = {}
        errors.append(f"House: {e}")

    data = {
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "year": year,
        "senate_source": senate_url,
        "house_source": "https://www.house.gov/legislative-activity",
        "senate_recess": senate_recess,
        "house_days": house_days,
        "errors": errors[:8],
    }
    save_chamber_calendar(data)
    return data


# Back-compat alias
pull_chamber_calendars = pull_session_calendar


def _monday_of_week(d):
    return d.fromordinal(d.toordinal() - d.weekday())


def _day_session_info(d, cache):
    recess = cache.get("senate_recess") or []
    house_days = cache.get("house_days") or {}
    iso = d.isoformat()
    s_status, s_label = _senate_status_for_day(d, recess)
    h_info = house_days.get(iso)
    if h_info:
        h_status, h_label = h_info["status"], h_info["label"]
    elif d.weekday() >= 5:
        h_status, h_label = "recess", "Weekend"
    else:
        h_status, h_label = "unknown", "No House data"
    return {
        "date": iso,
        "weekday": d.strftime("%a"),
        "label": d.strftime("%b %d"),
        "day_num": d.day,
        "is_today": d == date.today(),
        "senate": {"status": s_status, "label": s_label},
        "house": {"status": h_status, "label": h_label},
    }


def build_session_calendar_week(week_start=None, cache=None):
    if cache is None:
        cache = load_chamber_calendar()
    if week_start:
        try:
            monday = _monday_of_week(date.fromisoformat(week_start))
        except ValueError:
            monday = _monday_of_week(date.today())
    else:
        monday = _monday_of_week(date.today())

    days = []
    for i in range(7):
        d = monday.fromordinal(monday.toordinal() + i)
        days.append(_day_session_info(d, cache))

    return {
        "view": "week",
        "week_start": monday.isoformat(),
        "week_end": monday.fromordinal(monday.toordinal() + 6).isoformat(),
        "prev_week": monday.fromordinal(monday.toordinal() - 7).isoformat(),
        "next_week": monday.fromordinal(monday.toordinal() + 7).isoformat(),
        "days": days,
        "year": cache.get("year", monday.year),
    }


def build_session_calendar_month(year=None, month=None, cache=None):
    if cache is None:
        cache = load_chamber_calendar()
    today = date.today()
    year = int(year or cache.get("year") or today.year)
    month = int(month or today.month)
    if month < 1 or month > 12:
        month = today.month

    first = date(year, month, 1)
    if month == 12:
        last = date(year, 12, 31)
    else:
        last = date(year, month + 1, 1).fromordinal(date(year, month + 1, 1).toordinal() - 1)

    cal = calendar.Calendar(firstweekday=6)  # Sunday-first month grid
    weeks = []
    for week in cal.monthdatescalendar(year, month):
        row = []
        for d in week:
            if d.month != month:
                row.append({
                    "in_month": False,
                    "date": d.isoformat(),
                    "day_num": d.day,
                })
            else:
                info = _day_session_info(d, cache)
                info["in_month"] = True
                row.append(info)
        weeks.append(row)

    prev_m, prev_y = (month - 1, year) if month > 1 else (12, year - 1)
    next_m, next_y = (month + 1, year) if month < 12 else (1, year + 1)

    return {
        "view": "month",
        "year": year,
        "month": month,
        "month_name": first.strftime("%B %Y"),
        "month_start": first.isoformat(),
        "month_end": last.isoformat(),
        "prev_month": f"{prev_y}-{prev_m:02d}",
        "next_month": f"{next_y}-{next_m:02d}",
        "weeks": weeks,
        "weekday_headers": ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"],
    }


def build_calendar_week(week_start=None, events=None, tracked_only=False):
    """Legacy name — session week view."""
    return build_session_calendar_week(week_start=week_start)


def build_committee_ongoings(hearings, govtrack_cache=None):
    """
    Group hearings and metadata by tracked committee for the on-goings page.
    Returns list of dicts sorted by label.
    """
    if govtrack_cache is None:
        govtrack_cache = load_govtrack_cache()
    gt = govtrack_cache.get("committees") or {}
    cfg = load_congress_config()
    today = date.today()
    horizon = (today.toordinal() + 21)  # upcoming window ~3 weeks
    news_cutoff = today.toordinal() - 30

    sections = []
    for entry in cfg.get("committees") or []:
        label = entry.get("label", "")
        if not label:
            continue
        comm_hearings = [h for h in hearings if h.get("committee") == label]

        def _sort_key(h):
            try:
                return date.fromisoformat(h.get("date", "9999"))
            except Exception:
                return date.max

        upcoming = []
        recent = []
        news = []
        for h in comm_hearings:
            try:
                d_ord = date.fromisoformat(h.get("date", today.isoformat())).toordinal()
            except Exception:
                d_ord = today.toordinal()
            src = h.get("source", "")
            st = h.get("status", "")
            if st == "Upcoming" and d_ord >= today.toordinal() and d_ord <= horizon:
                upcoming.append(h)
            elif d_ord >= today.toordinal() - 14:
                recent.append(h)
            if src == "rss" and d_ord >= news_cutoff:
                news.append(h)

        upcoming.sort(key=_sort_key)
        recent.sort(key=_sort_key, reverse=True)
        news.sort(key=_sort_key, reverse=True)

        meta = gt.get(label) or {}
        sections.append({
            "label": label,
            "chamber": entry.get("chamber", ""),
            "system_code": entry.get("system_code", ""),
            "govtrack_code": entry.get("govtrack_code", ""),
            "meta": meta,
            "upcoming": upcoming[:8],
            "recent": recent[:6],
            "news": news[:5],
            "total": len(comm_hearings),
        })

    return sections

# ── Social / X Feed Pulling ───────────────────────────────────────────────────

def _strip_tweet_html(text):
    """Remove embedded blockquote/script tags from rss.app Twitter descriptions."""
    text = re.sub(r"<script[^>]*>.*?</script>", "", text, flags=re.DOTALL)
    text = re.sub(r"<blockquote[^>]*>", "", text)
    text = re.sub(r"</blockquote>", "", text)
    text = clean_html(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def pull_social_feeds(social_feeds=None, silent=False):
    """
    Fetch X / Twitter rss.app feeds and store items in social_feed.json.
    Returns a list of newly added items.
    """
    if social_feeds is None:
        social_feeds = load_social_feeds()

    existing_items = load_social_items()
    existing_urls  = {item["url"] for item in existing_items}
    cutoff = date.today().toordinal() - RSS_CUTOFF_DAYS
    new_items = []

    for feed in social_feeds:
        if not feed.get("active", True) or not feed.get("url", "").strip():
            continue
        if not silent:
            print(f"  Social feed: {feed['name']} ... ", end="", flush=True)
        added = 0
        try:
            raw_items = fetch_feed_raw(feed["url"])
            for item in raw_items:
                url = item.get("url", "")
                if not url or url in existing_urls:
                    continue

                item_date_str = parse_date_str(item["date"])
                try:
                    if date.fromisoformat(item_date_str).toordinal() < cutoff:
                        continue
                except Exception:
                    pass

                title   = clean_html(item.get("title", "")).strip()
                summary = _strip_tweet_html(item.get("summary", ""))
                if not title:
                    title = summary[:120]

                # Skip pure retweet noise with no text
                if not title and not summary:
                    continue

                social_item = {
                    "url":       url,
                    "title":     title,
                    "summary":   summary,
                    "date":      item_date_str,
                    "committee": feed.get("committee", "Other"),
                    "feed_name": feed["name"],
                    "author":    item.get("author", ""),
                    "fetched":   datetime.now().strftime("%Y-%m-%d"),
                }
                new_items.append(social_item)
                existing_urls.add(url)
                added += 1

            feed["last_fetched"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            feed["last_status"]  = "OK"
            if not silent:
                print(f"{added} new item(s)")
        except Exception as e:
            if not silent:
                print(f"FAILED ({e})")
            feed["last_status"] = f"Error: {e}"

    if new_items:
        combined = new_items + existing_items
        # Keep most recent SOCIAL_LIMIT items
        combined.sort(key=lambda x: x.get("date", ""), reverse=True)
        save_social_items(combined[:SOCIAL_LIMIT])

    save_social_feeds(social_feeds)
    return new_items

# ── Congress.gov API ──────────────────────────────────────────────────────────

CONGRESS_CONFIG_FILE = _data_path("congress_config.json")
CONGRESS_API_BASE    = "https://api.congress.gov/v3"
CONGRESS_API_CURRENT = 119
CONGRESS_API_LIMIT_PER_COMMITTEE = 20
CONGRESS_API_LIST_LIMIT = 15       # per chamber (API committee= filter is unreliable)
CONGRESS_API_DELAY = 0.05          # ~20 req/s max during a pull

# Default committee systemCode → label (overridden by congress_config.json if present)
_DEFAULT_CONGRESS_COMMITTEES = [
    {"system_code": "ssas00", "label": "Senate Armed Services (SASC)"},
    {"system_code": "ssfr00", "label": "Senate Foreign Relations (SFRC)"},
    {"system_code": "slin00", "label": "Senate Intelligence (SSCI)"},
    {"system_code": "ssga00", "label": "Senate Homeland Security (SHSGAC)"},
    {"system_code": "hsas00", "label": "House Armed Services (HASC)"},
    {"system_code": "hsfa00", "label": "House Foreign Affairs (HFAC)"},
    {"system_code": "hlig00", "label": "House Intelligence (HPSCI)"},
    {"system_code": "hshm00", "label": "House Homeland Security (HHSC)"},
]


def load_congress_config():
    """API endpoints + tracked committee system codes for Congress.gov pulls."""
    if os.path.exists(CONGRESS_CONFIG_FILE):
        with open(CONGRESS_CONFIG_FILE) as f:
            cfg = json.load(f)
    else:
        cfg = {
            "congress": CONGRESS_API_CURRENT,
            "api_base": CONGRESS_API_BASE,
            "meetings_limit_per_chamber": CONGRESS_API_LIST_LIMIT,
            "chambers": ["house", "senate"],
            "committees": _DEFAULT_CONGRESS_COMMITTEES,
        }
    committees = cfg.get("committees") or _DEFAULT_CONGRESS_COMMITTEES
    code_map = {
        c["system_code"]: c["label"]
        for c in committees
        if c.get("system_code") and c.get("label")
    }
    return {
        "congress": int(cfg.get("congress", CONGRESS_API_CURRENT)),
        "api_base": cfg.get("api_base", CONGRESS_API_BASE).rstrip("/"),
        "limit_per_committee": min(
            int(cfg.get("meetings_limit_per_committee", CONGRESS_API_LIMIT_PER_COMMITTEE)),
            50,
        ),
        "meetings_limit_per_chamber": min(
            int(cfg.get("meetings_limit_per_chamber", CONGRESS_API_LIST_LIMIT)),
            50,
        ),
        "chambers": cfg.get("chambers") or ["house", "senate"],
        "committee_codes": code_map,
        "committees": committees,
    }


def _committee_from_system_code(system_code, committee_codes):
    """Match full committee or subcommittee code (e.g. hsfa16 → hsfa00)."""
    if not system_code:
        return None
    if system_code in committee_codes:
        return committee_codes[system_code]
    if len(system_code) >= 4:
        parent = system_code[:4] + "00"
        if parent in committee_codes:
            return committee_codes[parent]
    return None


# Back-compat alias used elsewhere
CONGRESS_API_COMMITTEES = load_congress_config()["committee_codes"]

def fetch_json(url, api_key):
    """GET a URL from the Congress.gov API and return parsed JSON."""
    sep = "&" if "?" in url else "?"
    full_url = f"{url}{sep}api_key={api_key}&format=json"
    headers = {"User-Agent": "Mozilla/5.0 (Jamestown Foundation Hearing Tracker)"}
    req = Request(full_url, headers=headers)
    ctx = _make_ssl_context()
    with urlopen(req, timeout=12, context=ctx) as resp:
        return json.loads(resp.read())


def _chamber_for_system_code(system_code):
    if system_code.startswith("h"):
        return "house"
    if system_code.startswith("s"):
        return "senate"
    return None


def _parse_meeting_date_from_api(meeting):
    """Return (YYYY-MM-DD, is_tentative). API often omits dates on Senate records."""
    for d in meeting.get("dates") or []:
        raw = (d.get("date") or "")[:10]
        if raw and raw[0].isdigit():
            return raw, False
    for cont in meeting.get("meetingContinuations") or []:
        raw = (cont.get("date") or "")[:10]
        if raw and raw[0].isdigit():
            return raw, False
    upd = meeting.get("updateDate") or ""
    if upd:
        return str(upd)[:10], True
    return date.today().isoformat(), True


def _status_from_api(meeting, meeting_date):
    api_status = meeting.get("status") or "Scheduled"
    status = {
        "Scheduled":   "Upcoming",
        "Canceled":    "Cancelled",
        "Postponed":   "Postponed",
        "Rescheduled": "Upcoming",
    }.get(api_status, "Upcoming")
    if status == "Upcoming" and meeting_date < date.today().isoformat():
        status = "Completed"
    return status


def _meeting_notes_from_api(meeting):
    notes_parts = []
    for cont in meeting.get("meetingContinuations") or []:
        loc = cont.get("location") or {}
        room = loc.get("room", "")
        building = loc.get("building", "")
        if room or building:
            notes_parts.append(f"{room} {building}".strip())
        break
    for bill in (meeting.get("relatedItems") or {}).get("bills", []):
        num = bill.get("number", "")
        btype = bill.get("type", "")
        if num and btype:
            notes_parts.append(f"{btype} {num}")
    return "; ".join(notes_parts)


def _resolve_tracked_committee(meeting, committee_codes):
    """Return app committee label from API meeting committees, or None."""
    for c in meeting.get("committees") or []:
        name = _committee_from_system_code(c.get("systemCode", ""), committee_codes)
        if name:
            return name
    return None


def _find_hearing_for_api(hearings, event_id, committee_name, title):
    for h in hearings:
        if str(h.get("congress_event_id") or "") == event_id:
            if h.get("committee") == committee_name:
                return h
    title_lower = title.lower().strip()
    for h in hearings:
        if h.get("committee") != committee_name:
            continue
        ht = (h.get("topic") or "").lower().strip()
        if ht == title_lower or title_lower in ht or ht in title_lower:
            return h
    return None


def _apply_api_fields(hearing, meeting, *, committee_name, chamber, congress,
                      event_id, meeting_date, date_tentative):
    title = (meeting.get("title") or "").strip()
    if not title:
        title = f"{committee_name} – {meeting.get('type') or 'Meeting'}"
    witnesses = "; ".join(
        w.get("name", "") for w in meeting.get("witnesses", []) if w.get("name")
    )
    notes = _meeting_notes_from_api(meeting)
    if date_tentative:
        notes = (notes + "; " if notes else "") + "Congress.gov: meeting date not in API (using update date)"

    hearing["committee"] = committee_name
    hearing["topic"] = title
    hearing["date"] = meeting_date
    hearing["witnesses"] = witnesses
    hearing["status"] = _status_from_api(meeting, meeting_date)
    hearing["angle"] = detect_angle(title)
    hearing["notes"] = notes.strip("; ")
    hearing["url"] = _congress_api_meeting_url(meeting, congress, chamber, event_id)
    hearing["congress_event_id"] = event_id
    hearing["source"] = "api"


def _fetch_api_meeting_detail(api_base, congress, chamber, event_id, detail_url,
                              api_key, cache):
    """One detail fetch per event_id per pull (Congress list repeats the same ids)."""
    if event_id in cache:
        return cache[event_id]
    url = detail_url or (
        f"{api_base}/committee-meeting/{congress}/{chamber}/{event_id}?format=json"
    )
    time.sleep(CONGRESS_API_DELAY)
    detail = fetch_json(url, api_key)
    meeting = detail.get("committeeMeeting", {})
    cache[event_id] = meeting
    return meeting


def _hearing_needs_api_reconcile(hearing):
    """Only re-fetch when the stored link is missing or uses a broken legacy path."""
    if hearing.get("source") != "api":
        return False
    if not hearing.get("congress_event_id"):
        return False
    url = hearing.get("url") or ""
    if not url:
        return True
    if "/committee-meeting/" in url:
        return True
    if "senate.gov/isvp" in url:
        return True
    return False


def pull_congress_api(hearings, api_key, silent=False, persist=True,
                      meeting_cache=None):
    """
    Pull recent committee meetings from Congress.gov (one list per chamber).
    Detail payloads are cached by event_id because the API returns the same
    recent meetings regardless of committee= query param.
    """
    empty = {"new": [], "updated": 0}
    if not api_key or not api_key.strip():
        if not silent:
            print("  Congress.gov API: no key configured — skipping.")
        return empty

    cfg = load_congress_config()
    congress = cfg["congress"]
    api_base = cfg["api_base"]
    list_limit = min(
        int(cfg.get("meetings_limit_per_chamber", CONGRESS_API_LIST_LIMIT)),
        50,
    )
    cutoff = date.today().toordinal() - RSS_CUTOFF_DAYS
    codes = cfg["committee_codes"]
    cache = meeting_cache if meeting_cache is not None else {}
    # Senate schedule XML already fills tracked Senate meetings; skip duplicate API work.
    schedule_event_ids = {
        str(h.get("congress_event_id"))
        for h in hearings
        if h.get("source") == "schedule" and h.get("congress_event_id")
    }

    new_items = []
    updated_count = 0
    changed = False
    seen_list_ids = set()

    if not silent:
        print("  Congress.gov API ... ", end="", flush=True)

    for chamber in cfg.get("chambers") or ["house", "senate"]:
        try:
            list_url = (
                f"{api_base}/committee-meeting/{congress}/{chamber}"
                f"?limit={list_limit}&sort=updateDate+desc"
            )
            data = fetch_json(list_url, api_key)
            meetings = data.get("committeeMeetings", [])
        except Exception:
            continue

        for mtg in meetings:
            event_id = str(mtg.get("eventId", ""))
            if not event_id or event_id in seen_list_ids:
                continue
            seen_list_ids.add(event_id)
            if chamber == "senate" and event_id in schedule_event_ids:
                continue

            try:
                meeting = _fetch_api_meeting_detail(
                    api_base, congress, chamber, event_id,
                    mtg.get("url", ""), api_key, cache,
                )
            except Exception:
                continue

            resolved_committee = _resolve_tracked_committee(meeting, codes)
            if not resolved_committee:
                continue

            meeting_date, date_tentative = _parse_meeting_date_from_api(meeting)
            try:
                if date.fromisoformat(meeting_date).toordinal() < cutoff:
                    continue
            except Exception:
                pass

            title = (meeting.get("title") or "").strip()
            if not title:
                title = (
                    f"{resolved_committee} – {meeting.get('type') or 'Meeting'}"
                )

            existing = _find_hearing_for_api(
                hearings, event_id, resolved_committee, title
            )
            if existing:
                _apply_api_fields(
                    existing, meeting,
                    committee_name=resolved_committee,
                    chamber=chamber,
                    congress=congress,
                    event_id=event_id,
                    meeting_date=meeting_date,
                    date_tentative=date_tentative,
                )
                updated_count += 1
                changed = True
                continue

            h = {
                "id": next_id(hearings + new_items),
                "action": "Monitor only",
                "questions": "",
                "created": datetime.now().strftime("%Y-%m-%d"),
            }
            _apply_api_fields(
                h, meeting,
                committee_name=resolved_committee,
                chamber=chamber,
                congress=congress,
                event_id=event_id,
                meeting_date=meeting_date,
                date_tentative=date_tentative,
            )
            new_items.append(h)
            hearings.append(h)
            changed = True

    stale = [h for h in hearings if _hearing_needs_api_reconcile(h)]
    if stale:
        reconcile_api_hearings(
            hearings, api_key, cfg, silent=silent,
            meeting_cache=cache, persist=False,
        )
        changed = True

    if not silent:
        print(f"{len(new_items)} new, {updated_count} updated")

    if changed and persist:
        save_data(hearings)
    return {"new": new_items, "updated": updated_count}


def reconcile_api_hearings(hearings, api_key, cfg, silent=False,
                           meeting_cache=None, persist=True):
    """
    Re-fetch only API rows that still have broken legacy Congress.gov URLs.
    """
    congress = cfg["congress"]
    api_base = cfg["api_base"]
    codes = cfg["committee_codes"]
    cache = meeting_cache if meeting_cache is not None else {}
    changed = 0

    for h in hearings:
        if not _hearing_needs_api_reconcile(h):
            continue
        event_id = str(h.get("congress_event_id") or "").strip()
        chamber = _chamber_slug(
            "senate" if "Senate" in (h.get("committee") or "") else "house"
        )
        try:
            meeting = _fetch_api_meeting_detail(
                api_base, congress, chamber, event_id, "",
                api_key, cache,
            )
        except Exception:
            continue

        resolved = _resolve_tracked_committee(meeting, codes)
        new_url = _congress_api_meeting_url(meeting, congress, chamber, event_id)
        title = (meeting.get("title") or "").strip()

        if resolved and resolved != h.get("committee"):
            h["committee"] = resolved
            changed += 1
        if title and title != (h.get("topic") or ""):
            h["topic"] = title
            changed += 1
        if new_url and new_url != (h.get("url") or ""):
            h["url"] = new_url
            changed += 1
        if not resolved:
            note = "Congress.gov: meeting is not for a tracked committee"
            if note not in (h.get("notes") or ""):
                h["notes"] = ((h.get("notes") or "") + "; " + note).strip("; ")
            if h.get("committee") in _committee_site_urls(cfg):
                h["committee"] = "Other"
                changed += 1

    if changed and persist:
        save_data(hearings)
        if not silent:
            print(f"  Congress.gov API reconcile: {changed} record(s) corrected")
    return changed

# ── Federal Register API ──────────────────────────────────────────────────────

FR_DOCUMENTS_FILE = _data_path("fr_documents.json")
FR_WATCHLIST_FILE   = _data_path("fr_watchlist.json")
FR_API_BASE         = "https://www.federalregister.gov/api/v1"
FR_CUTOFF_DAYS      = 180
FR_PER_PAGE         = 100
FR_REQUEST_DELAY    = 0.12   # stay under ~10 req/s

FR_DOCUMENT_TYPES = ["PRORULE", "RULE", "NOTICE"]
FR_WORKFLOW_STATUSES = [
    "Watching", "Drafting", "Ready to file", "Filed", "Closed",
]

DEFAULT_FR_WATCHLIST = [
    {
        "name": "Sanctions / OFAC",
        "search_term": "sanctions",
        "document_types": ["PRORULE", "NOTICE"],
        "active": True,
    },
    {
        "name": "Export controls",
        "search_term": "export control",
        "document_types": ["PRORULE", "NOTICE"],
        "active": True,
    },
    {
        "name": "China / Indo-Pacific",
        "search_term": "China",
        "document_types": ["PRORULE", "NOTICE"],
        "active": True,
    },
    {
        "name": "Russia / Eurasia",
        "search_term": "Russia",
        "document_types": ["PRORULE", "NOTICE"],
        "active": True,
    },
    {
        "name": "Human rights",
        "search_term": "human rights",
        "document_types": ["PRORULE", "NOTICE"],
        "active": True,
    },
    {
        "name": "Cyber / information",
        "search_term": "cyber",
        "document_types": ["PRORULE", "NOTICE"],
        "active": True,
    },
]


def load_fr_documents():
    if os.path.exists(FR_DOCUMENTS_FILE):
        with open(FR_DOCUMENTS_FILE) as f:
            return json.load(f)
    return []


def save_fr_documents(documents):
    with open(FR_DOCUMENTS_FILE, "w") as f:
        json.dump(documents, f, indent=2)


def load_fr_watchlist():
    if os.path.exists(FR_WATCHLIST_FILE):
        with open(FR_WATCHLIST_FILE) as f:
            return json.load(f)
    save_fr_watchlist(DEFAULT_FR_WATCHLIST)
    return [dict(w) for w in DEFAULT_FR_WATCHLIST]


def save_fr_watchlist(watchlist):
    with open(FR_WATCHLIST_FILE, "w") as f:
        json.dump(watchlist, f, indent=2)


def fr_next_id(documents):
    return max((d["id"] for d in documents), default=0) + 1


def fr_agency_names(doc):
    agencies = doc.get("agencies") or []
    return [a.get("name", "") for a in agencies if a.get("name")]


def fr_comment_period_label(comments_close_on):
    """Human-readable comment window from comments_close_on (YYYY-MM-DD)."""
    if not comments_close_on:
        return "No deadline listed"
    try:
        close = date.fromisoformat(comments_close_on[:10])
    except Exception:
        return "Unknown"
    today = date.today()
    if close < today:
        return "Closed"
    days = (close - today).days
    if days == 0:
        return "Closes today"
    if days <= 7:
        return f"Closes in {days}d"
    if days <= 14:
        return f"Closes in {days}d"
    return "Open"


def fr_regulations_url(docket_ids):
    if docket_ids:
        return f"https://www.regulations.gov/docket/{docket_ids[0]}"
    return ""


def fr_fetch_documents(search_term, document_types, per_page=FR_PER_PAGE):
    """Query Federal Register API; returns list of document dicts from API."""
    params = [
        ("per_page", str(per_page)),
        ("page", "1"),
        ("order", "newest"),
        ("conditions[term]", search_term.strip()),
    ]
    for dtype in document_types or FR_DOCUMENT_TYPES:
        params.append(("conditions[type][]", dtype))
    url = f"{FR_API_BASE}/documents.json?{urlencode(params)}"
    headers = {"User-Agent": "Mozilla/5.0 (Jamestown Foundation Hearing Tracker)"}
    req = Request(url, headers=headers)
    ctx = _make_ssl_context()
    with urlopen(req, timeout=30, context=ctx) as resp:
        data = json.loads(resp.read())
    return data.get("results") or []


def _fr_doc_from_api(raw, watch_name):
    """Map Federal Register API document to stored record."""
    title = (raw.get("title") or "").strip()
    abstract = (raw.get("abstract") or "").strip()
    agencies = fr_agency_names(raw)
    docket_ids = raw.get("docket_ids") or []
    if isinstance(docket_ids, str):
        docket_ids = [docket_ids] if docket_ids else []
    comments_close = raw.get("comments_close_on")
    if comments_close:
        comments_close = str(comments_close)[:10]
    pub = raw.get("publication_date") or ""
    if pub:
        pub = str(pub)[:10]
    text_for_angle = f"{title} {abstract} {' '.join(agencies)}"
    return {
        "document_number": raw.get("document_number", ""),
        "title": title,
        "abstract": abstract[:2000] if abstract else "",
        "publication_date": pub,
        "comments_close_on": comments_close,
        "type": raw.get("type") or "",
        "agencies": agencies,
        "html_url": raw.get("html_url") or "",
        "docket_ids": docket_ids,
        "regulations_url": fr_regulations_url(docket_ids),
        "angle": detect_angle(text_for_angle),
        "workflow_status": "Watching",
        "draft_comment": "",
        "notes": "",
        "watch_name": watch_name,
        "source": "federal_register_api",
    }


def pull_federal_register(documents, watchlist, silent=False):
    """
    Pull documents for each active watchlist search term.
    Returns list of newly added document dicts (with id assigned).
    """
    new_items = []
    cutoff = date.today().toordinal() - FR_CUTOFF_DAYS
    by_number = {d["document_number"]: d for d in documents if d.get("document_number")}

    for watch in watchlist:
        if not watch.get("active", True):
            continue
        term = (watch.get("search_term") or "").strip()
        if not term:
            continue
        name = watch.get("name") or term
        dtypes = watch.get("document_types") or FR_DOCUMENT_TYPES
        if not silent:
            print(f"  Federal Register: {name} ({term!r}) ... ", end="", flush=True)
        try:
            time.sleep(FR_REQUEST_DELAY)
            results = fr_fetch_documents(term, dtypes)
            added = 0
            updated = 0
            for raw in results:
                dn = raw.get("document_number")
                if not dn:
                    continue
                pub = raw.get("publication_date") or ""
                try:
                    if pub and date.fromisoformat(str(pub)[:10]).toordinal() < cutoff:
                        continue
                except Exception:
                    pass
                mapped = _fr_doc_from_api(raw, name)
                if dn in by_number:
                    existing = by_number[dn]
                    if mapped.get("comments_close_on"):
                        existing["comments_close_on"] = mapped["comments_close_on"]
                        updated += 1
                    continue
                mapped["id"] = fr_next_id(documents + new_items)
                mapped["created"] = datetime.now().strftime("%Y-%m-%d")
                new_items.append(mapped)
                documents.append(mapped)
                by_number[dn] = mapped
                added += 1
            watch["last_fetched"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            watch["last_status"] = "OK"
            if not silent:
                up = f", {updated} updated" if updated else ""
                print(f"{added} new{up}")
        except URLError as e:
            watch["last_status"] = f"Error: {e.reason}"
            if not silent:
                print(f"FAILED ({e.reason})")
        except Exception as e:
            watch["last_status"] = f"Error: {e}"
            if not silent:
                print(f"FAILED ({e})")

    save_fr_documents(documents)
    save_fr_watchlist(watchlist)
    return new_items

# ── CRUD ──────────────────────────────────────────────────────────────────────

def add_hearing(hearings):
    header("ADD HEARING")
    h = {
        "id":        next_id(hearings),
        "date":      ask_date("Date"),
        "committee": pick("Committee:", COMMITTEES),
        "topic":     ask("Hearing topic"),
        "witnesses": ask("Witnesses (optional)", required=False),
        "angle":     pick("Jamestown angle:", JAMESTOWN_ANGLES),
        "action":    pick("Recommended action:", ACTIONS),
        "status":    pick("Status:", STATUS_OPTIONS),
        "url":       ask("Link/URL (optional)", required=False),
        "source":    "manual",
        "created":   datetime.now().strftime("%Y-%m-%d"),
    }
    print("\n  Questions to ask (one per line, blank to finish):")
    questions = []
    while True:
        q = input("  Q: ").strip()
        if not q:
            break
        questions.append(q)
    h["questions"] = " | ".join(questions)
    h["notes"]     = ask("Additional notes (optional)", required=False)
    hearings.append(h)
    save_data(hearings)
    print(f"\n  Hearing #{h['id']:03d} added.")

def list_hearings(hearings, filter_status=None):
    if not hearings:
        print("\n  No hearings tracked yet.")
        return
    filtered = sorted(
        [h for h in hearings if not filter_status or h.get("status") == filter_status],
        key=lambda x: x["date"]
    )
    print(f"\n  Showing {len(filtered)} hearing(s):")
    divider()
    for h in filtered:
        format_hearing(h)

def view_hearing(hearings):
    hid = input("\n  Enter hearing ID: ").strip()
    if not hid.isdigit():
        print("  Invalid ID.")
        return
    matches = [h for h in hearings if h["id"] == int(hid)]
    if not matches:
        print("  Hearing not found.")
        return
    format_hearing(matches[0], verbose=True)

def edit_hearing(hearings):
    hid = input("\n  Enter hearing ID to edit: ").strip()
    if not hid.isdigit():
        print("  Invalid ID.")
        return
    matches = [h for h in hearings if h["id"] == int(hid)]
    if not matches:
        print("  Hearing not found.")
        return
    h = matches[0]
    format_hearing(h)
    print("  Press Enter to keep current value.\n")

    new_date = ask_date(f"Date [{h['date']}]", required=False)
    if new_date:
        h["date"] = new_date

    print(f"\n  Current committee: {h['committee']}")
    if input("  Change committee? (y/N): ").strip().lower() == "y":
        h["committee"] = pick("Committee:", COMMITTEES)

    new_topic = ask(f"Topic [{h['topic']}]", required=False)
    if new_topic:
        h["topic"] = new_topic

    new_w = ask(f"Witnesses [{h.get('witnesses', '')}]", required=False)
    if new_w:
        h["witnesses"] = new_w

    print(f"\n  Current angle: {h.get('angle', '--')}")
    if input("  Change angle? (y/N): ").strip().lower() == "y":
        h["angle"] = pick("Jamestown angle:", JAMESTOWN_ANGLES)

    print(f"\n  Current action: {h.get('action', '--')}")
    if input("  Change action? (y/N): ").strip().lower() == "y":
        h["action"] = pick("Recommended action:", ACTIONS)

    print(f"\n  Current status: {h.get('status', '--')}")
    if input("  Change status? (y/N): ").strip().lower() == "y":
        h["status"] = pick("Status:", STATUS_OPTIONS)

    if input("\n  Update questions? (y/N): ").strip().lower() == "y":
        print("  Enter questions (blank to finish):")
        questions = []
        while True:
            q = input("  Q: ").strip()
            if not q:
                break
            questions.append(q)
        h["questions"] = " | ".join(questions)

    new_notes = ask(f"Notes [{h.get('notes', '')}]", required=False)
    if new_notes:
        h["notes"] = new_notes

    save_data(hearings)
    print(f"\n  Hearing #{h['id']:03d} updated.")

def delete_hearing(hearings):
    hid = input("\n  Enter hearing ID to delete: ").strip()
    if not hid.isdigit():
        print("  Invalid ID.")
        return
    matches = [h for h in hearings if h["id"] == int(hid)]
    if not matches:
        print("  Hearing not found.")
        return
    format_hearing(matches[0])
    if input("  Delete this hearing? (yes/N): ").strip().lower() == "yes":
        hearings[:] = [h for h in hearings if h["id"] != int(hid)]
        save_data(hearings)
        print("  Deleted.")
    else:
        print("  Cancelled.")

def search_hearings(hearings):
    term = input("\n  Search term: ").strip().lower()
    if not term:
        return
    results = sorted(
        [h for h in hearings if any(
            term in h.get(f, "").lower()
            for f in ["topic", "committee", "angle", "witnesses", "notes"]
        )],
        key=lambda x: x["date"]
    )
    print(f"\n  Found {len(results)} match(es):")
    divider()
    for h in results:
        format_hearing(h)

def upcoming_summary(hearings):
    header("NEXT 14 DAYS")
    today_ord  = date.today().toordinal()
    cutoff_ord = today_ord + 14
    upcoming = sorted(
        [h for h in hearings
         if h.get("status") in ("Upcoming", None)
         and today_ord <= date.fromisoformat(h["date"]).toordinal() <= cutoff_ord],
        key=lambda x: x["date"]
    )
    if not upcoming:
        print("\n  No upcoming hearings in the next 14 days.")
    else:
        for h in upcoming:
            format_hearing(h, verbose=True)

def manage_feeds(hearings, feeds):
    while True:
        header("RSS FEED MANAGEMENT")
        for i, f in enumerate(feeds, 1):
            status     = "ON " if f.get("active", True) else "OFF"
            last_fetch = f.get("last_fetched", "never")
            last_ok    = f.get("last_status", "--")
            print(f"  {i:2}. [{status}] {f['name']}")
            print(f"        Last: {last_fetch} | {last_ok}")
        divider()
        print("  A. Pull all feeds now")
        print("  T. Toggle feed on/off")
        print("  N. Add custom feed")
        print("  D. Delete feed")
        print("  B. Back")
        divider()
        choice = input("  Select: ").strip().upper()

        if choice == "A":
            print()
            new_items = pull_rss_feeds(hearings, feeds)
            print(f"\n  Done. {len(new_items)} new hearing(s) added.")

        elif choice == "T":
            idx = input("  Feed number to toggle: ").strip()
            if idx.isdigit() and 1 <= int(idx) <= len(feeds):
                f = feeds[int(idx) - 1]
                f["active"] = not f.get("active", True)
                save_feeds(feeds)
                print(f"  {f['name']} {'enabled' if f['active'] else 'disabled'}.")
            else:
                print("  Invalid selection.")

        elif choice == "N":
            print()
            url  = ask("Feed URL")
            name = ask("Display name")
            comm = pick("Map to committee:", COMMITTEES + ["Multiple"])
            feeds.append({"name": name, "url": url, "committee": comm, "active": True})
            save_feeds(feeds)
            print(f"  Feed '{name}' added.")

        elif choice == "D":
            idx = input("  Feed number to delete: ").strip()
            if idx.isdigit() and 1 <= int(idx) <= len(feeds):
                removed = feeds.pop(int(idx) - 1)
                save_feeds(feeds)
                print(f"  Removed '{removed['name']}'.")
            else:
                print("  Invalid selection.")

        elif choice == "B":
            break

def export_to_excel(hearings):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  Run: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Hearing Tracker"

    headers = ["ID", "Date", "Days Away", "Status", "Source", "Committee",
               "Topic", "Witnesses", "Angle", "Action", "Questions", "Link", "Notes"]
    hdr_fill = PatternFill("solid", start_color="1F3864")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill("solid", start_color="DCE6F1")
    rss_fill = PatternFill("solid", start_color="E2EFDA")

    for row_num, h in enumerate(sorted(hearings, key=lambda x: x["date"]), 2):
        try:
            delta = (date.fromisoformat(h["date"]) - date.today()).days
            days_away = "Today" if delta == 0 else (f"In {delta}d" if delta > 0 else f"{abs(delta)}d ago")
        except Exception:
            days_away = ""

        is_rss    = h.get("source") == "rss"
        questions = h.get("questions", "").replace(" | ", "\n")
        row = [h["id"], h["date"], days_away, h.get("status", ""),
               "RSS" if is_rss else "Manual",
               h["committee"], h["topic"], h.get("witnesses", ""),
               h.get("angle", ""), h.get("action", ""), questions,
               h.get("url", ""), h.get("notes", "")]

        fill = rss_fill if is_rss else (alt_fill if row_num % 2 == 0 else None)
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    widths = [5, 12, 10, 12, 8, 30, 42, 28, 22, 18, 40, 35, 35]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    filename = f"jamestown_hearings_{date.today()}.xlsx"
    wb.save(filename)
    print(f"\n  Exported to {filename}")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n+========================================================================+")
    print("|        JAMESTOWN FOUNDATION -- CONGRESSIONAL HEARING TRACKER          |")
    print("+========================================================================+")

    if not HAS_FEEDPARSER:
        print("\n  NOTE: feedparser not installed -- using fallback XML parser.")
        print("        For better results: pip install feedparser\n")

    hearings = load_data()
    feeds    = load_feeds()

    # Auto-pull on first run
    if all(not f.get("last_fetched") for f in feeds):
        print("\n  First run -- pulling RSS feeds automatically...")
        new_items = pull_rss_feeds(hearings, feeds)
        print(f"  {len(new_items)} hearing(s) imported.\n")

    while True:
        total     = len(hearings)
        upcoming  = sum(1 for h in hearings
                        if h.get("status") in ("Upcoming", None)
                        and date.fromisoformat(h["date"]) >= date.today())
        rss_count = sum(1 for h in hearings if h.get("source") == "rss")

        print(f"\n  Tracked: {total}  |  Upcoming: {upcoming}  |  From RSS: {rss_count}  |  {date.today()}")
        divider()
        print("   1. Pull RSS feeds now")
        print("   2. Add hearing manually")
        print("   3. List all hearings")
        print("   4. Upcoming (next 14 days)")
        print("   5. View hearing details")
        print("   6. Edit hearing")
        print("   7. Delete hearing")
        print("   8. Search hearings")
        print("   9. Filter by status")
        print("  10. Manage RSS feeds")
        print("  11. Export to Excel")
        print("   0. Quit")
        divider()

        choice = input("  Select: ").strip()

        if choice == "1":
            print()
            new_items = pull_rss_feeds(hearings, feeds)
            print(f"\n  Done. {len(new_items)} new hearing(s) added.")
        elif choice == "2":
            add_hearing(hearings)
        elif choice == "3":
            list_hearings(hearings)
        elif choice == "4":
            upcoming_summary(hearings)
        elif choice == "5":
            view_hearing(hearings)
        elif choice == "6":
            edit_hearing(hearings)
        elif choice == "7":
            delete_hearing(hearings)
        elif choice == "8":
            search_hearings(hearings)
        elif choice == "9":
            status = pick("Filter by status:", STATUS_OPTIONS)
            list_hearings(hearings, filter_status=status)
        elif choice == "10":
            manage_feeds(hearings, feeds)
        elif choice == "11":
            export_to_excel(hearings)
        elif choice == "0":
            print("\n  Goodbye.\n")
            sys.exit(0)
        else:
            print("  Invalid choice.")

# ── DC Location / Heatmap helpers ─────────────────────────────────────────────

# Known Capitol Hill building coordinates (lat, lng)
DC_BUILDINGS = {
    # Senate side
    "russell":  (38.8928, -77.0073),
    "dirksen":  (38.8924, -77.0064),
    "hart":     (38.8921, -77.0058),
    "capitol":  (38.8899, -77.0091),
    # House side
    "cannon":   (38.8872, -77.0076),
    "longworth":(38.8863, -77.0086),
    "rayburn":  (38.8851, -77.0096),
    # Other
    "ford":     (38.8878, -77.0058),   # Ford House Office Building
    "o'neill":  (38.8875, -77.0062),   # Thomas P. O'Neill Jr. Federal Building
    "cvc":      (38.8895, -77.0083),   # Capitol Visitor Center
}

# Keyword → building key (checked in order against notes/topic, case-insensitive)
_BUILDING_KEYWORDS = [
    ("russell",   "russell"),
    ("dirksen",   "dirksen"),
    ("hart",      "hart"),
    ("cannon",    "cannon"),
    ("longworth", "longworth"),
    ("rayburn",   "rayburn"),
    ("ford",      "ford"),
    ("o'neill",   "o'neill"),
    ("capitol visitor", "cvc"),
    ("webex",     None),  # virtual — skip
]

# Committee → default building when no explicit location is found
_COMMITTEE_DEFAULT_BUILDING = {
    "Senate Armed Services (SASC)":       "russell",
    "Senate Foreign Relations (SFRC)":    "dirksen",
    "Senate Intelligence (SSCI)":         "hart",
    "Senate Homeland Security (SHSGAC)":  "dirksen",
    "House Armed Services (HASC)":        "rayburn",
    "House Foreign Affairs (HFAC)":       "longworth",
    "House Intelligence (HPSCI)":         "cvc",
    "House Homeland Security (HHSC)":     "cannon",
}

def resolve_hearing_location(h):
    """
    Return (lat, lng, building_label) for a hearing dict, or None if unknown.
    Checks the 'notes' field first (Congress.gov API populates room/building),
    then falls back to the committee's primary building.
    """
    text = (h.get("notes") or "").lower()

    for keyword, bkey in _BUILDING_KEYWORDS:
        if keyword in text:
            if bkey is None:
                return None   # virtual meeting
            coords = DC_BUILDINGS.get(bkey)
            if coords:
                label = keyword.title() + " Building"
                return (*coords, label)

    # Fall back to committee default
    bkey = _COMMITTEE_DEFAULT_BUILDING.get(h.get("committee", ""))
    if bkey:
        coords = DC_BUILDINGS[bkey]
        return (*coords, bkey.title() + " (default)")

    return None

def build_heatmap_points(hearings):
    """
    Return a list of dicts suitable for JSON serialisation and Leaflet.heat.
    Each point: {lat, lng, weight, label, committee, topic, date, url}
    Weight is boosted for action-required / upcoming hearings.
    """
    points = []
    for h in hearings:
        loc = resolve_hearing_location(h)
        if not loc:
            continue
        lat, lng, building = loc

        # Weight: upcoming + action-required hearings stand out more
        weight = 0.4
        if h.get("status") == "Upcoming":
            weight = 0.7
        if h.get("action") in ("Send brief", "Send questions",
                               "Offer testimony", "Request meeting"):
            weight = 1.0

        points.append({
            "lat":       lat,
            "lng":       lng,
            "weight":    weight,
            "building":  building,
            "committee": h.get("committee", ""),
            "topic":     h.get("topic", ""),
            "date":      h.get("date", ""),
            "status":    h.get("status", ""),
            "url":       h.get("url", ""),
            "id":        h.get("id"),
        })
    return points

if __name__ == "__main__":
    main()

